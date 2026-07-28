# Sistema: Gerador de Apostas por Padrões - Dia de Sorte
# Desenvolvido para: Márcio Fernando Maia
# Funcionalidade: Geração COMPLETA de todas as combinações possíveis para cada padrão
# Versão: 2.0 - Com persistência em banco, viabilidade e busca inteligente

from itertools import combinations, product
from math import comb
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO
import time

# Importações do modelo Sorteio
try:
    from models.sorteio import Sorteio, db
except ImportError:
    from models.sorteio import Sorteio
    try:
        from app import db
    except ImportError:
        from flask_sqlalchemy import SQLAlchemy
        db = SQLAlchemy()

# Importações dos repositórios (SQL raw - SEM SQLAlchemy ORM para evitar conflitos)
try:
    from models.gerador_padroes import (
        PadroesRepository, CombinacoesRepository, SorteiosReaisRepository,
        gerar_hash_dezenas
    )
    REPOSITORIOS_DISPONIVEIS = True
except ImportError:
    PadroesRepository = None
    CombinacoesRepository = None
    SorteiosReaisRepository = None
    gerar_hash_dezenas = None
    REPOSITORIOS_DISPONIVEIS = False

# Aliases para compatibilidade (não são mais classes ORM)
MODELOS_DISPONIVEIS = REPOSITORIOS_DISPONIVEIS


class GeradorPadroesCompletoService:
    """
    Serviço para análise e geração COMPLETA de apostas baseadas em padrões de dezenas.

    Padrão: Representa a distribuição de números por faixa (coluna).
    Exemplo: "0 0 0 0 0 2 3" = 5 números da faixa 0, 1 da faixa 2, 1 da faixa 3

    Faixas:
    - 0: 01-09 (9 números disponíveis)
    - 1: 10-19 (10 números disponíveis)
    - 2: 20-29 (10 números disponíveis)
    - 3: 30-31 (2 números disponíveis)
    """

    # Configuração das faixas (SISTEMA ANTIGO - PARA COMPATIBILIDADE)
    FAIXAS = {
        0: list(range(1, 10)),    # 01-09 (9 números)
        1: list(range(10, 20)),   # 10-19 (10 números)
        2: list(range(20, 30)),   # 20-29 (10 números)
        3: [30, 31]               # 30-31 (2 números)
    }

    TAMANHOS_FAIXAS = {
        0: 9,
        1: 10,
        2: 10,
        3: 2
    }

    # 🆕 CONFIGURAÇÃO BAIXAS/MÉDIAS/ALTAS (NOVO SISTEMA)
    # Padrão: "0 0 1 1 2 2 2" significa 2 baixas, 2 médias, 3 altas
    FAIXAS_BMA = {
        0: list(range(1, 11)),    # BAIXAS: 01-10 (10 números)
        1: list(range(11, 21)),   # MÉDIAS: 11-20 (10 números)
        2: list(range(21, 32))    # ALTAS: 21-31 (11 números)
    }

    TAMANHOS_FAIXAS_BMA = {
        0: 10,  # Baixas
        1: 10,  # Médias
        2: 11   # Altas
    }

    # =========================================================================
    # CÁLCULO DE PADRÕES E JOGOS POSSÍVEIS
    # =========================================================================

    @staticmethod
    def _padrao_string_para_contagem(padrao_str):
        """
        Converte string de padrão para contagem por faixa.

        Exemplo: "0 0 0 0 0 2 3" -> {0: 5, 1: 0, 2: 1, 3: 1}
        """
        partes = padrao_str.strip().split()
        contagem = {0: 0, 1: 0, 2: 0, 3: 0}

        for p in partes:
            try:
                faixa = int(p)
                if faixa in contagem:
                    contagem[faixa] += 1
            except ValueError:
                continue

        return contagem

    @staticmethod
    def _contagem_para_padrao_string(contagem):
        """
        Converte contagem por faixa para string de padrão ordenado.

        Exemplo: {0: 5, 1: 0, 2: 1, 3: 1} -> "0 0 0 0 0 2 3"
        """
        partes = []
        for faixa in sorted(contagem.keys()):
            partes.extend([str(faixa)] * contagem[faixa])
        return ' '.join(partes)

    @staticmethod
    def calcular_jogos_possiveis(padrao_str):
        """
        Calcula quantos jogos (combinações) são possíveis para um padrão.

        Usa combinatória: C(n, k) para cada faixa e multiplica os resultados.

        Exemplo: "0 0 0 0 0 2 3"
        - Faixa 0: C(9, 5) = 126
        - Faixa 1: C(10, 0) = 1
        - Faixa 2: C(10, 1) = 10
        - Faixa 3: C(2, 1) = 2
        - Total: 126 × 1 × 10 × 2 = 2.520

        Returns:
            int: Número total de jogos possíveis (0 se padrão inválido)
        """
        contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str)

        # Verificar se soma 7
        if sum(contagem.values()) != 7:
            return 0

        total = 1
        for faixa, qtd in contagem.items():
            tamanho_faixa = GeradorPadroesCompletoService.TAMANHOS_FAIXAS[faixa]

            # Verificar se é possível escolher qtd números dessa faixa
            if qtd > tamanho_faixa:
                return 0

            total *= comb(tamanho_faixa, qtd)

        return total

    @staticmethod
    def verificar_padrao_viavel(padrao_str):
        """
        Verifica se um padrão é viável (pode gerar pelo menos 1 jogo).
        """
        return GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str) > 0

    # =========================================================================
    # LISTAGEM DE TODOS OS PADRÕES POSSÍVEIS
    # =========================================================================

    @staticmethod
    def listar_todos_padroes():
        """
        Lista TODOS os padrões possíveis (viáveis) com suas estatísticas.

        Gera todas as combinações de 7 elementos usando faixas 0, 1, 2, 3,
        filtrando apenas os viáveis.

        Returns:
            list: Lista de dicts com {padrao, jogos_possiveis, contagem}
        """
        padroes = []

        # Gerar todas as combinações possíveis de contagens que somam 7
        # Faixa 0: 0-7 (max 9, mas limitado por outras faixas)
        # Faixa 1: 0-7 (max 10)
        # Faixa 2: 0-7 (max 10)
        # Faixa 3: 0-2 (max 2)

        for f0 in range(min(8, 10)):  # 0-7 da faixa 0 (max 9 disponíveis)
            for f1 in range(min(8 - f0, 11)):  # restante para faixa 1
                for f2 in range(min(8 - f0 - f1, 11)):  # restante para faixa 2
                    f3 = 7 - f0 - f1 - f2  # o que sobra vai para faixa 3

                    # Verificar limites
                    if f3 < 0 or f3 > 2:
                        continue
                    if f0 > 9:
                        continue
                    if f1 > 10:
                        continue
                    if f2 > 10:
                        continue

                    contagem = {0: f0, 1: f1, 2: f2, 3: f3}
                    padrao_str = GeradorPadroesCompletoService._contagem_para_padrao_string(contagem)
                    jogos = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)

                    if jogos > 0:
                        padroes.append({
                            'padrao': padrao_str,
                            'jogos_possiveis': jogos,
                            'contagem': contagem,
                            'descricao': f"F0:{f0} | F1:{f1} | F2:{f2} | F3:{f3}"
                        })

        # Ordenar por quantidade de jogos possíveis (decrescente)
        padroes.sort(key=lambda x: x['jogos_possiveis'], reverse=True)

        return padroes

    @staticmethod
    def listar_padroes_com_historico():
        """
        Lista todos os padrões viáveis COM informações do histórico de sorteios.

        Adiciona:
        - frequencia: quantas vezes o padrão apareceu nos sorteios
        - atraso: há quantos concursos não sai
        - ultimo_concurso: último concurso onde apareceu
        - status: 'frequente', 'atrasado', 'faltante'

        Returns:
            dict com:
            - padroes: lista completa de padrões
            - total_padroes: quantidade total de padrões
            - total_jogos: soma de todos os jogos possíveis
            - estatisticas: resumo geral
        """
        # Obter todos os padrões viáveis
        padroes = GeradorPadroesCompletoService.listar_todos_padroes()

        # Obter histórico de sorteios
        try:
            sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).all()
        except Exception:
            sorteios = []

        # Contar padrões nos sorteios
        contagem_historico = Counter()
        ultimo_concurso_padrao = {}

        ultimo_concurso = sorteios[0].concurso if sorteios else 0
        ultimo_resultado = GeradorPadroesCompletoService.obter_ultimo_resultado_padrao(sorteio=sorteios[0] if sorteios else None)
        padrao_ultimo_sorteio = ultimo_resultado['padrao'] if ultimo_resultado else None

        for sorteio in sorteios:
            numeros = sorteio.get_posicoes_lista()
            padrao = GeradorPadroesCompletoService._numeros_para_padrao(numeros)

            contagem_historico[padrao] += 1

            if padrao not in ultimo_concurso_padrao:
                ultimo_concurso_padrao[padrao] = sorteio.concurso

        # Enriquecer padrões com dados do histórico
        total_jogos = 0
        padroes_frequentes = 0
        padroes_atrasados = 0
        padroes_faltantes = 0

        for p in padroes:
            padrao_str = p['padrao']
            frequencia = contagem_historico.get(padrao_str, 0)
            ultimo_conc = ultimo_concurso_padrao.get(padrao_str, None)

            p['frequencia'] = frequencia
            p['ultimo_concurso'] = ultimo_conc
            p['eh_padrao_ultimo_concurso'] = (padrao_str == padrao_ultimo_sorteio)

            if frequencia == 0:
                p['atraso'] = None
                p['status'] = 'faltante'
                padroes_faltantes += 1
            else:
                p['atraso'] = ultimo_concurso - ultimo_conc
                if p['atraso'] > 50:
                    p['status'] = 'atrasado'
                    padroes_atrasados += 1
                else:
                    p['status'] = 'frequente'
                    padroes_frequentes += 1

            total_jogos += p['jogos_possiveis']

        total_sorteios = len(sorteios)
        top_frequencia = GeradorPadroesCompletoService.obter_top_frequencia(
            padroes, limite=3, total_sorteios=total_sorteios
        )

        return {
            'padroes': padroes,
            'total_padroes': len(padroes),
            'total_jogos_possiveis': total_jogos,
            'ultimo_concurso': ultimo_concurso,
            'total_sorteios_analisados': total_sorteios,
            'ultimo_resultado': ultimo_resultado,
            'top_frequencia': top_frequencia,
            'padrao_mais_frequente': top_frequencia[0] if top_frequencia else None,
            'estatisticas': {
                'frequentes': padroes_frequentes,
                'atrasados': padroes_atrasados,
                'faltantes': padroes_faltantes
            }
        }

    @staticmethod
    def obter_top_frequencia(padroes, limite=3, total_sorteios=0):
        """
        Padrões que mais apareceram no histórico (frequência decrescente).
        Inclui percentual sobre o total de concursos analisados.
        """
        com_freq = [p for p in padroes if p.get('frequencia', 0) > 0]
        ordenados = sorted(
            com_freq,
            key=lambda x: (-x['frequencia'], -(x.get('ultimo_concurso') or 0))
        )[:limite]

        resultado = []
        for pos, p in enumerate(ordenados, start=1):
            freq = p['frequencia']
            pct = round((freq / total_sorteios) * 100, 2) if total_sorteios else 0
            resultado.append({
                'posicao': pos,
                'padrao': p['padrao'],
                'descricao': p.get('descricao', ''),
                'frequencia': freq,
                'percentual_concursos': pct,
                'ultimo_concurso': p.get('ultimo_concurso'),
                'atraso': p.get('atraso'),
                'status': p.get('status', ''),
                'jogos_possiveis': p.get('jogos_possiveis', 0),
            })
        return resultado

    @staticmethod
    def obter_top_frequencia_elite(limite=3):
        """
        Top N padrões posicionais mais frequentes no histórico,
        com concursos de exemplo para o Simulador Elite (modo 3+7).
        """
        try:
            sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).all()
        except Exception:
            sorteios = []

        if not sorteios:
            return {
                'total_sorteios_analisados': 0,
                'top_frequencia': [],
            }

        contagem = Counter()
        concursos_por_padrao = defaultdict(list)

        for sorteio in sorteios:
            numeros = sorteio.get_posicoes_lista()
            if not numeros or len(numeros) < 7:
                continue

            padrao = GeradorPadroesCompletoService._numeros_para_padrao(numeros)
            contagem[padrao] += 1

            data_fmt = ''
            if getattr(sorteio, 'data_sorteio', None):
                try:
                    data_fmt = sorteio.data_sorteio.strftime('%d/%m/%Y')
                except Exception:
                    data_fmt = str(sorteio.data_sorteio)

            concursos_por_padrao[padrao].append({
                'concurso': sorteio.concurso,
                'data': data_fmt,
                'numeros': [f'{n:02d}' for n in sorted(numeros)],
            })

        total_sorteios = len(sorteios)
        resultado = []
        for pos, (padrao, freq) in enumerate(contagem.most_common(limite), start=1):
            f0, f1, f2, f3 = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao)
            pct = round((freq / total_sorteios) * 100, 2) if total_sorteios else 0
            concursos = concursos_por_padrao.get(padrao, [])
            ultimo_concurso = concursos[0]['concurso'] if concursos else None
            atraso = (sorteios[0].concurso - ultimo_concurso) if ultimo_concurso else None

            resultado.append({
                'posicao': pos,
                'padrao': padrao,
                'descricao': 'F0:{} | F1:{} | F2:{} | F3:{}'.format(f0, f1, f2, f3),
                'frequencia': freq,
                'percentual_concursos': pct,
                'ultimo_concurso': ultimo_concurso,
                'atraso': atraso,
                'concursos': concursos,
            })

        return {
            'total_sorteios_analisados': total_sorteios,
            'top_frequencia': resultado,
        }

    @staticmethod
    def obter_ultimo_resultado_padrao(sorteio=None):
        """
        Retorna o último concurso do banco com dezenas e padrão calculado
        (faixas 0-3 por dezena) para comparação na lista de padrões.
        """
        try:
            if sorteio is None:
                sorteio = Sorteio.query.order_by(Sorteio.concurso.desc()).first()
        except Exception:
            return None

        if not sorteio:
            return None

        numeros = sorteio.get_posicoes_lista()
        if not numeros or len(numeros) < 7:
            return None

        padrao = GeradorPadroesCompletoService._numeros_para_padrao(numeros)
        contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao)
        f0, f1, f2, f3 = contagem[0], contagem[1], contagem[2], contagem[3]

        data_fmt = None
        if getattr(sorteio, 'data_sorteio', None):
            try:
                data_fmt = sorteio.data_sorteio.strftime('%d/%m/%Y')
            except Exception:
                data_fmt = str(sorteio.data_sorteio)

        return {
            'concurso': sorteio.concurso,
            'data_sorteio': data_fmt,
            'mes_sorte': sorteio.mes_sorte,
            'numeros': numeros,
            'numeros_formatados': [str(n).zfill(2) for n in numeros],
            'padrao': padrao,
            'descricao': 'F0:{} | F1:{} | F2:{} | F3:{}'.format(f0, f1, f2, f3),
            'distribuicao': {'faixa_0': f0, 'faixa_1': f1, 'faixa_2': f2, 'faixa_3': f3},
            'legenda_faixas': '0=01-09 | 1=10-19 | 2=20-29 | 3=30-31',
        }

    @staticmethod
    def _numeros_para_padrao(numeros):
        """
        Converte lista de 7 números para string de padrão.

        Exemplo: [1, 5, 7, 12, 25, 28, 31] -> "0 0 0 1 2 2 3"
        """
        faixas = []
        for num in sorted(numeros):
            if num <= 9:
                faixas.append('0')
            elif num <= 19:
                faixas.append('1')
            elif num <= 29:
                faixas.append('2')
            else:
                faixas.append('3')
        return ' '.join(faixas)

    # =========================================================================
    # GERAÇÃO COMPLETA DE TODAS AS COMBINAÇÕES
    # =========================================================================

    @staticmethod
    def gerar_todas_combinacoes(padrao_str):
        """
        Gera TODAS as combinações possíveis para um padrão específico.

        ATENÇÃO: Pode gerar milhões de combinações para alguns padrões!
        Use com cuidado e implemente paginação/streaming no frontend.

        Args:
            padrao_str: String do padrão (ex: "0 0 0 0 0 2 3")

        Returns:
            dict com:
            - padrao: o padrão solicitado
            - total: quantidade de jogos gerados
            - jogos: lista de todas as combinações
            - tempo_geracao: tempo em segundos
        """
        import time
        inicio = time.time()

        contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str)

        # Verificar viabilidade
        jogos_esperados = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
        if jogos_esperados == 0:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': 'Padrão inválido ou inviável',
                'tempo_geracao': 0
            }

        # Gerar combinações para cada faixa
        combinacoes_por_faixa = []

        for faixa in range(4):
            qtd = contagem[faixa]
            if qtd > 0:
                numeros_faixa = GeradorPadroesCompletoService.FAIXAS[faixa]
                combos = list(combinations(numeros_faixa, qtd))
                combinacoes_por_faixa.append(combos)
            else:
                combinacoes_por_faixa.append([tuple()])  # Tupla vazia para manter o product

        # Combinar todas as faixas usando produto cartesiano
        jogos = []
        for combo in product(*combinacoes_por_faixa):
            # Flatten e ordenar
            jogo = []
            for parte in combo:
                jogo.extend(parte)
            jogo.sort()
            jogos.append(jogo)

        tempo = time.time() - inicio

        return {
            'padrao': padrao_str,
            'contagem': contagem,
            'total': len(jogos),
            'total_esperado': jogos_esperados,
            'jogos': jogos,
            'tempo_geracao': round(tempo, 3)
        }

    @staticmethod
    def gerar_combinacoes_paginado(padrao_str, pagina=1, por_pagina=100):
        """
        Gera combinações de forma paginada para melhor performance.

        Args:
            padrao_str: String do padrão
            pagina: Número da página (1-indexed)
            por_pagina: Quantidade de jogos por página

        Returns:
            dict com jogos da página solicitada
        """
        import time
        inicio = time.time()

        contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str)

        # Verificar viabilidade
        total_jogos = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
        if total_jogos == 0:
            return {
                'padrao': padrao_str,
                'pagina': pagina,
                'por_pagina': por_pagina,
                'total': 0,
                'total_paginas': 0,
                'jogos': [],
                'erro': 'Padrão inválido ou inviável'
            }

        total_paginas = (total_jogos + por_pagina - 1) // por_pagina

        if pagina < 1:
            pagina = 1
        if pagina > total_paginas:
            pagina = total_paginas

        # Gerar combinações para cada faixa
        combinacoes_por_faixa = []

        for faixa in range(4):
            qtd = contagem[faixa]
            if qtd > 0:
                numeros_faixa = GeradorPadroesCompletoService.FAIXAS[faixa]
                combos = list(combinations(numeros_faixa, qtd))
                combinacoes_por_faixa.append(combos)
            else:
                combinacoes_por_faixa.append([tuple()])

        # Calcular índices
        inicio_idx = (pagina - 1) * por_pagina
        fim_idx = min(inicio_idx + por_pagina, total_jogos)

        # Gerar apenas os jogos necessários usando iteração
        jogos = []
        idx = 0

        for combo in product(*combinacoes_por_faixa):
            if idx >= fim_idx:
                break
            if idx >= inicio_idx:
                jogo = []
                for parte in combo:
                    jogo.extend(parte)
                jogo.sort()
                jogos.append(jogo)
            idx += 1

        tempo = time.time() - inicio

        return {
            'padrao': padrao_str,
            'contagem': contagem,
            'pagina': pagina,
            'por_pagina': por_pagina,
            'total': total_jogos,
            'total_paginas': total_paginas,
            'jogos': jogos,
            'jogos_na_pagina': len(jogos),
            'inicio': inicio_idx + 1,
            'fim': inicio_idx + len(jogos),
            'tempo_geracao': round(tempo, 3)
        }

    @staticmethod
    def gerar_combinacoes_generator(padrao_str):
        """
        Gera combinações usando generator para economia de memória.
        Útil para streaming/exportação de grandes quantidades.

        Yields:
            list: Cada jogo como lista de 7 números ordenados
        """
        contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str)

        # Verificar viabilidade
        if GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str) == 0:
            return

        # Gerar combinações para cada faixa
        combinacoes_por_faixa = []

        for faixa in range(4):
            qtd = contagem[faixa]
            if qtd > 0:
                numeros_faixa = GeradorPadroesCompletoService.FAIXAS[faixa]
                combos = list(combinations(numeros_faixa, qtd))
                combinacoes_por_faixa.append(combos)
            else:
                combinacoes_por_faixa.append([tuple()])

        # Usar product como generator
        for combo in product(*combinacoes_por_faixa):
            jogo = []
            for parte in combo:
                jogo.extend(parte)
            jogo.sort()
            yield jogo

    # =========================================================================
    # 🆕 GERAÇÃO COM DEZENAS SELECIONADAS (BAIXAS/MÉDIAS/ALTAS)
    # =========================================================================

    @staticmethod
    def gerar_com_dezenas_selecionadas(padrao_str, dezenas_selecionadas):
        """
        Gera apostas usando APENAS as dezenas selecionadas pelo usuário,
        respeitando o padrão de baixas (01-10), médias (11-20) e altas (21-31).

        Padrão CORRIGIDO:
        - 0 = BAIXA (01-10)
        - 1 = MÉDIA (11-20)
        - 2 = ALTA (21-31)
        - 3 = EXTREMA (30-31) — compatibilidade com o padrão antigo

        Exemplo:
        Padrão "0 0 1 1 2 2 2" = 2 baixas, 2 médias, 3 altas
        Dezenas [1, 4, 11, 18, 21, 22, 29] -> gera jogo respeitando o padrão

        Args:
            padrao_str: String do padrão (ex: "0 0 1 1 2 2 2")
            dezenas_selecionadas: Lista de dezenas escolhidas pelo usuário

        Returns:
            dict com jogos gerados usando apenas as dezenas selecionadas
        """
        import time
        inicio = time.time()

        # Validar dezenas selecionadas
        if not dezenas_selecionadas or len(dezenas_selecionadas) < 7:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': 'Selecione pelo menos 7 dezenas',
                'tempo_geracao': 0
            }

        # Validar padrão informado
        partes = [p.strip() for p in padrao_str.strip().split() if p.strip()]
        if len(partes) != 7:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão deve ter 7 posições, mas veio com {len(partes)}',
                'tempo_geracao': 0
            }

        invalidos = [p for p in partes if p not in {'0', '1', '2', '3'}]
        if invalidos:
            valores = ', '.join(sorted(set(invalidos)))
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão contém valores inválidos: {valores} (use apenas 0, 1, 2 ou 3)',
                'tempo_geracao': 0
            }

        # Separar dezenas em baixas, médias, altas e extremas (30-31)
        # Nova regra: altas FIXAS em 21-29; 30-31 só entram se houver faixa 3 no padrão
        possui_faixa_extrema = '3' in partes
        baixas_disponiveis = sorted([d for d in dezenas_selecionadas if 1 <= d <= 10])
        medias_disponiveis = sorted([d for d in dezenas_selecionadas if 11 <= d <= 20])
        altas_limite_superior = 29  # sempre 21-29
        altas_disponiveis = sorted([d for d in dezenas_selecionadas if 21 <= d <= altas_limite_superior])
        faixa_altas_rotulo = "21-29"
        extremas_disponiveis = sorted([d for d in dezenas_selecionadas if 30 <= d <= 31]) if possui_faixa_extrema else []

        # Contar quantas de cada faixa o padrão exige
        qtd_baixas = partes.count('0')
        qtd_medias = partes.count('1')
        qtd_altas = partes.count('2')
        qtd_extremas = partes.count('3')

        # Validar se temos dezenas suficientes
        if len(baixas_disponiveis) < qtd_baixas:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão exige {qtd_baixas} baixas (01-10), mas você selecionou apenas {len(baixas_disponiveis)}',
                'tempo_geracao': 0
            }

        if len(medias_disponiveis) < qtd_medias:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão exige {qtd_medias} médias (11-20), mas você selecionou apenas {len(medias_disponiveis)}',
                'tempo_geracao': 0
            }

        if len(altas_disponiveis) < qtd_altas:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão exige {qtd_altas} altas ({faixa_altas_rotulo}), mas você selecionou apenas {len(altas_disponiveis)}',
                'tempo_geracao': 0
            }

        if len(extremas_disponiveis) < qtd_extremas:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Padrão exige {qtd_extremas} extremas (30-31), mas você selecionou apenas {len(extremas_disponiveis)}',
                'tempo_geracao': 0
            }

        # Gerar todas as combinações
        jogos = []

        # Combinações de baixas
        if qtd_baixas > 0:
            combos_baixas = list(combinations(baixas_disponiveis, qtd_baixas))
        else:
            combos_baixas = [tuple()]

        # Combinações de médias
        if qtd_medias > 0:
            combos_medias = list(combinations(medias_disponiveis, qtd_medias))
        else:
            combos_medias = [tuple()]

        # Combinações de altas
        if qtd_altas > 0:
            combos_altas = list(combinations(altas_disponiveis, qtd_altas))
        else:
            combos_altas = [tuple()]

        # Combinações de extremas (faixa 3 - 30/31)
        if qtd_extremas > 0:
            combos_extremas = list(combinations(extremas_disponiveis, qtd_extremas))
        else:
            combos_extremas = [tuple()]

        # Produto cartesiano de todas as combinações
        anomalies = 0  # Contagem defensiva para garantir 7 dezenas
        for baixas, medias, altas, extremas in product(combos_baixas, combos_medias, combos_altas, combos_extremas):
            jogo = sorted(list(baixas) + list(medias) + list(altas) + list(extremas))
            if len(jogo) != 7:
                anomalies += 1
                continue
            jogos.append(jogo)

        tempo = time.time() - inicio

        if anomalies > 0:
            return {
                'padrao': padrao_str,
                'total': 0,
                'jogos': [],
                'erro': f'Falha ao gerar combinações: {anomalies} combinações retornaram tamanho diferente de 7',
                'tempo_geracao': round(tempo, 3)
            }

        return {
            'padrao': padrao_str,
            'total': len(jogos),
            'jogos': jogos,
            'tempo_geracao': round(tempo, 3),
            'dezenas_usadas': {
                'baixas': baixas_disponiveis,
                'medias': medias_disponiveis,
                'altas': altas_disponiveis,
                'extremas': extremas_disponiveis
            },
            'distribuicao_padrao': {
                'baixas': qtd_baixas,
                'medias': qtd_medias,
                'altas': qtd_altas,
                'extremas': qtd_extremas
            },
            'anomalies_descartadas': anomalies if anomalies > 0 else None
        }

    # =========================================================================
    # EXPORTAÇÃO
    # =========================================================================

    @staticmethod
    def exportar_para_txt(padrao_str, mes_abrev=''):
        """
        Gera conteúdo TXT para exportação no formato padrão do sistema.

        Formato: "01 05 07 12 25 28 31 Nov"

        Args:
            padrao_str: String do padrão
            mes_abrev: Abreviação do mês (ex: 'Nov')

        Returns:
            str: Conteúdo do arquivo TXT
        """
        linhas = []

        for jogo in GeradorPadroesCompletoService.gerar_combinacoes_generator(padrao_str):
            numeros_fmt = ' '.join(str(n).zfill(2) for n in jogo)
            if mes_abrev:
                linhas.append(f"{numeros_fmt} {mes_abrev}")
            else:
                linhas.append(numeros_fmt)

        return '\n'.join(linhas)

    @staticmethod
    def exportar_para_xlsx_data(padrao_str, mes_nome=''):
        """
        Gera dados estruturados para exportação XLSX.

        Returns:
            list: Lista de dicts prontos para DataFrame/Excel
        """
        dados = []

        for idx, jogo in enumerate(GeradorPadroesCompletoService.gerar_combinacoes_generator(padrao_str), 1):
            linha = {
                'Jogo': idx,
                'D1': jogo[0],
                'D2': jogo[1],
                'D3': jogo[2],
                'D4': jogo[3],
                'D5': jogo[4],
                'D6': jogo[5],
                'D7': jogo[6],
                'Padrão': padrao_str
            }
            if mes_nome:
                linha['Mês'] = mes_nome
            dados.append(linha)

        return dados

    # =========================================================================
    # BUSCA E FILTROS
    # =========================================================================

    @staticmethod
    def buscar_padrao(padrao_str):
        """
        Busca informações completas de um padrão específico.

        Returns:
            dict com todas as informações do padrão ou None se não encontrado
        """
        padroes = GeradorPadroesCompletoService.listar_padroes_com_historico()

        for p in padroes['padroes']:
            if p['padrao'] == padrao_str:
                return p

        # Se não encontrado na lista, verificar se é viável
        jogos = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
        if jogos > 0:
            return {
                'padrao': padrao_str,
                'jogos_possiveis': jogos,
                'contagem': GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str),
                'frequencia': 0,
                'atraso': None,
                'status': 'faltante'
            }

        return None

    @staticmethod
    def filtrar_padroes(status=None, min_jogos=None, max_jogos=None, ordenar_por='jogos_possiveis'):
        """
        Filtra padrões por critérios específicos.

        Args:
            status: 'frequente', 'atrasado', 'faltante' ou None para todos
            min_jogos: Mínimo de jogos possíveis
            max_jogos: Máximo de jogos possíveis
            ordenar_por: 'jogos_possiveis', 'frequencia', 'atraso'

        Returns:
            list: Lista de padrões filtrados
        """
        dados = GeradorPadroesCompletoService.listar_padroes_com_historico()
        padroes = dados['padroes']

        # Aplicar filtros
        if status:
            padroes = [p for p in padroes if p['status'] == status]

        if min_jogos is not None:
            padroes = [p for p in padroes if p['jogos_possiveis'] >= min_jogos]

        if max_jogos is not None:
            padroes = [p for p in padroes if p['jogos_possiveis'] <= max_jogos]

        # Ordenar
        reverse = True
        if ordenar_por == 'atraso':
            # Colocar None (faltantes) no final
            padroes = sorted(padroes, key=lambda x: (x['atraso'] is None, x['atraso'] or 0), reverse=True)
        elif ordenar_por == 'frequencia':
            padroes = sorted(padroes, key=lambda x: x['frequencia'], reverse=True)
        else:
            padroes = sorted(padroes, key=lambda x: x['jogos_possiveis'], reverse=True)

        return padroes

    # =========================================================================
    # ESTATÍSTICAS
    # =========================================================================

    @staticmethod
    def obter_estatisticas_gerais():
        """
        Retorna estatísticas gerais sobre todos os padrões.
        """
        dados = GeradorPadroesCompletoService.listar_padroes_com_historico()
        padroes = dados['padroes']

        if not padroes:
            return {}

        jogos_totais = sum(p['jogos_possiveis'] for p in padroes)
        frequencias = [p['frequencia'] for p in padroes]
        jogos_list = [p['jogos_possiveis'] for p in padroes]

        # Top 5 mais jogos
        top_jogos = sorted(padroes, key=lambda x: x['jogos_possiveis'], reverse=True)[:5]

        # Top 5 menos jogos
        bottom_jogos = sorted(padroes, key=lambda x: x['jogos_possiveis'])[:5]

        # Top 5 mais frequentes
        top_frequentes = sorted(padroes, key=lambda x: x['frequencia'], reverse=True)[:5]

        # Top 5 mais atrasados (excluindo faltantes)
        atrasados_validos = [p for p in padroes if p['atraso'] is not None]
        top_atrasados = sorted(atrasados_validos, key=lambda x: x['atraso'], reverse=True)[:5]

        return {
            'total_padroes': len(padroes),
            'total_jogos_possiveis': jogos_totais,
            'media_jogos_por_padrao': round(jogos_totais / len(padroes), 2),
            'min_jogos': min(jogos_list),
            'max_jogos': max(jogos_list),
            'estatisticas': dados['estatisticas'],
            'top_5_mais_jogos': top_jogos,
            'top_5_menos_jogos': bottom_jogos,
            'top_5_mais_frequentes': top_frequentes,
            'top_5_mais_atrasados': top_atrasados
        }

    # =========================================================================
    # CLASSIFICAÇÃO DE VIABILIDADE ESTATÍSTICA
    # =========================================================================

    @staticmethod
    def _obter_estatisticas_dezenas():
        """Obtém estatísticas de frequência e atraso das dezenas."""
        stats = {}
        try:
            sorteios = Sorteio.query.all()
            if not sorteios:
                return stats

            contagem = {i: 0 for i in range(1, 32)}
            ultima_ocorrencia = {i: 0 for i in range(1, 32)}
            ultimo_concurso = len(sorteios)

            for idx, sorteio in enumerate(sorteios, 1):
                dezenas = sorteio.get_posicoes_lista()
                for d in dezenas:
                    if d:
                        contagem[d] += 1
                        ultima_ocorrencia[d] = idx

            for d in range(1, 32):
                atraso = ultimo_concurso - ultima_ocorrencia[d]
                stats[d] = {
                    'frequencia': contagem[d],
                    'atraso': atraso
                }

        except Exception as e:
            pass

        return stats

    @staticmethod
    def avaliar_viabilidade_combinacao(dezenas, stats_dezenas=None):
        """
        Avalia a viabilidade estatística de uma combinação.

        Retorna:
            tuple: (viavel: bool, motivo: str ou None)
        """
        if stats_dezenas is None:
            stats_dezenas = GeradorPadroesCompletoService._obter_estatisticas_dezenas()

        if not stats_dezenas:
            return True, None

        # Regra 1: Verificar dezenas muito atrasadas (mais de 80 concursos)
        dezenas_muito_atrasadas = [
            d for d in dezenas
            if stats_dezenas.get(d, {}).get('atraso', 0) > 80
        ]
        if len(dezenas_muito_atrasadas) >= 3:
            return False, f"Muitas dezenas atrasadas: {dezenas_muito_atrasadas}"

        # Regra 2: Verificar soma fora do intervalo típico (70-170)
        soma = sum(dezenas)
        if soma < 70 or soma > 170:
            return False, f"Soma atípica: {soma} (fora do intervalo 70-170)"

        # Regra 3: Verificar sequências longas (5+ consecutivos)
        dezenas_ord = sorted(dezenas)
        max_seq = 1
        seq_atual = 1
        for i in range(1, len(dezenas_ord)):
            if dezenas_ord[i] == dezenas_ord[i - 1] + 1:
                seq_atual += 1
                max_seq = max(max_seq, seq_atual)
            else:
                seq_atual = 1

        if max_seq >= 5:
            return False, f"Sequência longa: {max_seq} consecutivos"

        # Regra 4: Verificar todas pares ou todas ímpares
        pares = len([d for d in dezenas if d % 2 == 0])
        if pares == 0 or pares == 7:
            return False, f"Todas {'pares' if pares == 7 else 'ímpares'}"

        return True, None

    @staticmethod
    def gerar_combinacoes_com_viabilidade(padrao_str, limite=None):
        """
        Gera combinações com classificação de viabilidade.

        Returns:
            list: Lista de dicts com {dezenas, soma, viavel, motivo_inviavel}
        """
        resultado = []
        stats_dezenas = GeradorPadroesCompletoService._obter_estatisticas_dezenas()

        count = 0
        for jogo in GeradorPadroesCompletoService.gerar_combinacoes_generator(padrao_str):
            viavel, motivo = GeradorPadroesCompletoService.avaliar_viabilidade_combinacao(
                jogo, stats_dezenas
            )

            resultado.append({
                'dezenas': jogo,
                'soma': sum(jogo),
                'viavel': viavel,
                'motivo_inviavel': motivo
            })

            count += 1
            if limite and count >= limite:
                break

        return resultado

    # =========================================================================
    # PERSISTÊNCIA NO BANCO DE DADOS
    # =========================================================================

    @staticmethod
    def salvar_padrao_no_banco(padrao_str, forcar_atualizacao=False):
        """
        Salva um padrão no banco de dados.
        Se já existir, retorna o registro existente (ou atualiza se forcar_atualizacao=True).
        """
        if not REPOSITORIOS_DISPONIVEIS or PadroesRepository is None:
            return None

        try:
            # Verificar se já existe
            padrao_existente = PadroesRepository.buscar_por_padrao(padrao_str)

            if padrao_existente:
                if forcar_atualizacao:
                    # Atualizar estatísticas
                    info = GeradorPadroesCompletoService.buscar_padrao(padrao_str)
                    if info:
                        PadroesRepository.atualizar(
                            padrao_str,
                            frequencia=info.get('frequencia', 0),
                            atraso=info.get('atraso'),
                            status=info.get('status', 'faltante')
                        )
                return padrao_existente

            # Criar novo padrão
            contagem = GeradorPadroesCompletoService._padrao_string_para_contagem(padrao_str)
            jogos = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
            info = GeradorPadroesCompletoService.buscar_padrao(padrao_str) or {}

            return PadroesRepository.inserir(
                padrao_str=padrao_str,
                descricao=f"0:{contagem[0]} | 1:{contagem[1]} | 2:{contagem[2]} | 3:{contagem[3]}",
                jogos_possiveis=jogos,
                frequencia=info.get('frequencia', 0),
                atraso=info.get('atraso'),
                status=info.get('status', 'faltante'),
                viavel=info.get('frequencia', 0) > 0
            )

        except Exception as e:
            return None

    @staticmethod
    def salvar_combinacoes_no_banco(padrao_str, limite=None):
        """
        Salva combinações geradas no banco de dados.
        Evita duplicatas usando hash das dezenas.

        Returns:
            dict: {salvas: int, existentes: int, erro: str ou None}
        """
        if not REPOSITORIOS_DISPONIVEIS or CombinacoesRepository is None:
            return {'salvas': 0, 'existentes': 0, 'erro': 'Repositórios não disponíveis'}

        try:
            # Garantir que o padrão existe
            padrao = GeradorPadroesCompletoService.salvar_padrao_no_banco(padrao_str)
            if not padrao:
                return {'salvas': 0, 'existentes': 0, 'erro': 'Erro ao salvar padrão'}

            padrao_id = padrao.get('id') if isinstance(padrao, dict) else padrao.id
            stats_dezenas = GeradorPadroesCompletoService._obter_estatisticas_dezenas()

            # Preparar dados em lote
            combinacoes_data = []
            count = 0

            for jogo in GeradorPadroesCompletoService.gerar_combinacoes_generator(padrao_str):
                # Avaliar viabilidade
                viavel, motivo = GeradorPadroesCompletoService.avaliar_viabilidade_combinacao(
                    jogo, stats_dezenas
                )

                combinacoes_data.append({
                    'dezenas': jogo,
                    'viavel': viavel,
                    'motivo_inviavel': motivo
                })

                count += 1
                if limite and count >= limite:
                    break

                # Inserir em lotes de 1000
                if len(combinacoes_data) >= 1000:
                    CombinacoesRepository.inserir_lote(padrao_id, combinacoes_data)
                    combinacoes_data = []

            # Inserir restante
            if combinacoes_data:
                CombinacoesRepository.inserir_lote(padrao_id, combinacoes_data)

            return {'salvas': count, 'existentes': 0, 'erro': None}

        except Exception as e:
            return {'salvas': 0, 'existentes': 0, 'erro': str(e)}

    @staticmethod
    def obter_combinacoes_do_banco(padrao_str, pagina=1, por_pagina=100, apenas_viaveis=None):
        """
        Obtém combinações do banco de dados com paginação.

        Args:
            padrao_str: String do padrão
            pagina: Número da página
            por_pagina: Itens por página
            apenas_viaveis: True para apenas viáveis, False para apenas não viáveis, None para todos

        Returns:
            dict: {combinacoes, total, pagina, por_pagina, do_banco}
        """
        if not REPOSITORIOS_DISPONIVEIS or CombinacoesRepository is None or PadroesRepository is None:
            # Fallback: gerar na hora sem persistência
            resultado = GeradorPadroesCompletoService.gerar_combinacoes_paginado(
                padrao_str, pagina, por_pagina
            )
            resultado['do_banco'] = False
            return resultado

        try:
            padrao = PadroesRepository.buscar_por_padrao(padrao_str)
            if not padrao:
                # Padrão não está no banco - gerar na hora
                resultado = GeradorPadroesCompletoService.gerar_combinacoes_paginado(
                    padrao_str, pagina, por_pagina
                )
                resultado['do_banco'] = False
                return resultado

            padrao_id = padrao.get('id')

            # Contar total
            total = CombinacoesRepository.contar(padrao_id, apenas_viaveis)
            total_paginas = (total + por_pagina - 1) // por_pagina if total > 0 else 1

            # Buscar combinações com paginação
            offset = (pagina - 1) * por_pagina
            combinacoes = CombinacoesRepository.buscar_por_padrao_id(
                padrao_id, offset, por_pagina, apenas_viaveis
            )

            return {
                'padrao': padrao_str,
                'pagina': pagina,
                'por_pagina': por_pagina,
                'total': total,
                'total_paginas': total_paginas,
                'combinacoes': combinacoes,
                'do_banco': True
            }

        except Exception as e:
            # Fallback
            resultado = GeradorPadroesCompletoService.gerar_combinacoes_paginado(
                padrao_str, pagina, por_pagina
            )
            resultado['do_banco'] = False
            resultado['erro'] = str(e)
            return resultado

    # =========================================================================
    # BUSCA INTELIGENTE PÓS-SORTEIO
    # =========================================================================

    @staticmethod
    def analisar_sorteio_real(dezenas, concurso=None, mes_sorte=None, salvar=True):
        """
        Analisa um resultado de sorteio real.

        Identifica:
        - O padrão das dezenas sorteadas
        - Se o padrão já existe no sistema
        - Combinações geradas que contêm as 7 dezenas (match total)
        - Estatísticas do padrão

        Args:
            dezenas: Lista de 7 dezenas sorteadas
            concurso: Número do concurso (opcional)
            mes_sorte: Mês da sorte 1-12 (opcional)
            salvar: Se deve salvar o sorteio no banco

        Returns:
            dict: Análise completa do sorteio
        """
        # Validar dezenas
        if len(dezenas) != 7:
            return {'erro': 'São necessárias exatamente 7 dezenas'}

        dezenas = sorted([int(d) for d in dezenas])

        # Validar intervalo
        for d in dezenas:
            if d < 1 or d > 31:
                return {'erro': f'Dezena {d} fora do intervalo (1-31)'}

        # Verificar duplicatas
        if len(set(dezenas)) != 7:
            return {'erro': 'Dezenas duplicadas não são permitidas'}

        # Calcular padrão
        padrao_str = GeradorPadroesCompletoService._numeros_para_padrao(dezenas)

        # Buscar informações do padrão
        info_padrao = GeradorPadroesCompletoService.buscar_padrao(padrao_str)

        resultado = {
            'dezenas': dezenas,
            'soma': sum(dezenas),
            'padrao': padrao_str,
            'descricao_padrao': info_padrao.get('descricao', '') if info_padrao else '',
            'jogos_possiveis_padrao': GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str),
            'estatisticas_padrao': {
                'frequencia': info_padrao.get('frequencia', 0) if info_padrao else 0,
                'atraso': info_padrao.get('atraso') if info_padrao else None,
                'status': info_padrao.get('status', 'faltante') if info_padrao else 'faltante'
            },
            'concurso': concurso,
            'mes_sorte': mes_sorte,
            'combinacao_encontrada': False,
            'combinacao_id': None,
            'combinacao_viavel': None,
            'matches_parciais': {
                'total_6_acertos': 0,
                'total_5_acertos': 0,
                'total_4_acertos': 0,
                'detalhes': []
            }
        }

        # Avaliar viabilidade das dezenas sorteadas
        viavel, motivo = GeradorPadroesCompletoService.avaliar_viabilidade_combinacao(dezenas)
        resultado['viavel'] = viavel
        resultado['motivo_viabilidade'] = motivo

        # Buscar combinação exata no banco
        hash_dezenas = '-'.join([str(d).zfill(2) for d in dezenas])

        if REPOSITORIOS_DISPONIVEIS and CombinacoesRepository is not None:
            try:
                comb_exata = CombinacoesRepository.buscar_por_hash(hash_dezenas)
                if comb_exata:
                    resultado['combinacao_encontrada'] = True
                    resultado['combinacao_id'] = comb_exata.get('id')
                    resultado['combinacao_viavel'] = comb_exata.get('viavel')
            except Exception:
                pass

        # Buscar matches parciais (6, 5, 4 acertos)
        if REPOSITORIOS_DISPONIVEIS and CombinacoesRepository is not None:
            try:
                matches = CombinacoesRepository.buscar_matches_parciais(dezenas, min_acertos=4, limite=50)
                resultado['matches_parciais'] = {
                    'total_6_acertos': matches.get('total_6', 0),
                    'total_5_acertos': matches.get('total_5', 0),
                    'total_4_acertos': matches.get('total_4', 0),
                    'detalhes': matches.get('detalhes', [])
                }
            except Exception:
                pass

        # Salvar sorteio real no banco (se solicitado e repositórios disponíveis)
        if salvar and REPOSITORIOS_DISPONIVEIS and SorteiosReaisRepository is not None:
            try:
                # Verificar se já existe
                existente = SorteiosReaisRepository.buscar_por_hash(hash_dezenas)
                if not existente:
                    # Buscar padrao_id
                    padrao_id = None
                    if PadroesRepository is not None:
                        padrao_db = PadroesRepository.buscar_por_padrao(padrao_str)
                        if padrao_db:
                            padrao_id = padrao_db.get('id')

                    sorteio = SorteiosReaisRepository.inserir(
                        concurso=concurso,
                        dezenas=dezenas,
                        mes_sorte=mes_sorte,
                        padrao_str=padrao_str,
                        padrao_id=padrao_id,
                        combinacao_encontrada=resultado['combinacao_encontrada'],
                        combinacao_id=resultado.get('combinacao_id')
                    )

                    if sorteio:
                        resultado['sorteio_salvo'] = True
                        resultado['sorteio_id'] = sorteio.get('id')
                    else:
                        resultado['sorteio_salvo'] = False
                else:
                    resultado['sorteio_salvo'] = False
                    resultado['sorteio_existente'] = True
                    resultado['sorteio_id'] = existente.get('id')

            except Exception as e:
                resultado['erro_salvar'] = str(e)

        return resultado

    @staticmethod
    def _buscar_matches_parciais(dezenas_sorteadas, min_acertos=4, limite=50):
        """
        Busca combinações com correspondência parcial.

        Returns:
            dict: {total_6_acertos, total_5_acertos, total_4_acertos, detalhes}
        """
        resultado = {
            'total_6_acertos': 0,
            'total_5_acertos': 0,
            'total_4_acertos': 0,
            'detalhes': []
        }

        if not REPOSITORIOS_DISPONIVEIS or CombinacoesRepository is None:
            return resultado

        try:
            # Usar método do repositório que já faz a busca otimizada
            matches = CombinacoesRepository.buscar_matches_parciais(
                dezenas_sorteadas, min_acertos, limite
            )

            resultado['total_6_acertos'] = matches.get('total_6', 0)
            resultado['total_5_acertos'] = matches.get('total_5', 0)
            resultado['total_4_acertos'] = matches.get('total_4', 0)
            resultado['detalhes'] = matches.get('detalhes', [])

        except Exception as e:
            resultado['erro'] = str(e)

        return resultado

    @staticmethod
    def obter_historico_sorteios_reais(limite=50):
        """
        Obtém histórico de sorteios reais analisados.

        Returns:
            list: Lista de sorteios com suas análises
        """
        if not REPOSITORIOS_DISPONIVEIS or SorteiosReaisRepository is None:
            return []

        try:
            sorteios = SorteiosReaisRepository.listar_ultimos(limite)
            return [
                {
                    'id': s.get('id'),
                    'concurso': s.get('concurso'),
                    'dezenas': s.get('dezenas', []),
                    'mes_sorte': s.get('mes_sorte'),
                    'padrao': s.get('padrao_str'),
                    'combinacao_encontrada': s.get('combinacao_encontrada', False),
                    'data_registro': s.get('data_registro').isoformat() if s.get('data_registro') else None
                }
                for s in sorteios
            ]
        except Exception:
            return []

    # =========================================================================
    # EXPORTAÇÃO XLSX COM VIABILIDADE
    # =========================================================================

    @staticmethod
    def exportar_para_xlsx_completo(padrao_str, mes_nome='', incluir_viabilidade=True):
        """
        Exporta combinações para XLSX com informações de viabilidade.

        Returns:
            BytesIO: Arquivo XLSX em memória
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Apostas"

        # Estilos
        header_fill = PatternFill(start_color="D4B31A", end_color="D4B31A", fill_type="solid")
        header_font = Font(bold=True, color="2D2606")
        viavel_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        nao_viavel_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalho
        headers = ['#', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'Soma', 'Padrão', 'Mês']
        if incluir_viabilidade:
            headers.extend(['Viável', 'Motivo'])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Dados
        stats_dezenas = GeradorPadroesCompletoService._obter_estatisticas_dezenas() if incluir_viabilidade else None

        for idx, jogo in enumerate(GeradorPadroesCompletoService.gerar_combinacoes_generator(padrao_str), 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx).border = thin_border

            for j, num in enumerate(jogo, 2):
                cell = ws.cell(row=row, column=j, value=num)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            ws.cell(row=row, column=9, value=sum(jogo)).border = thin_border
            ws.cell(row=row, column=10, value=padrao_str).border = thin_border
            ws.cell(row=row, column=11, value=mes_nome).border = thin_border

            if incluir_viabilidade:
                viavel, motivo = GeradorPadroesCompletoService.avaliar_viabilidade_combinacao(
                    jogo, stats_dezenas
                )

                cell_viavel = ws.cell(row=row, column=12, value='Sim' if viavel else 'Não')
                cell_viavel.alignment = Alignment(horizontal='center')
                cell_viavel.border = thin_border
                cell_viavel.fill = viavel_fill if viavel else nao_viavel_fill

                ws.cell(row=row, column=13, value=motivo or '').border = thin_border

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # =========================================================================
    # ESTATÍSTICAS DO BANCO
    # =========================================================================

    @staticmethod
    def obter_estatisticas_banco():
        """
        Obtém estatísticas do banco de dados.

        Returns:
            dict: Estatísticas de padrões, combinações e sorteios
        """
        stats = {
            'repositorios_disponiveis': REPOSITORIOS_DISPONIVEIS,
            'modelos_disponiveis': MODELOS_DISPONIVEIS,  # Mantém para compatibilidade
            'padroes': {'total': 0, 'frequentes': 0, 'atrasados': 0, 'faltantes': 0},
            'combinacoes': {'total': 0, 'viaveis': 0, 'nao_viaveis': 0},
            'sorteios_analisados': 0
        }

        if not REPOSITORIOS_DISPONIVEIS:
            return stats

        try:
            if PadroesRepository is not None:
                stats['padroes']['total'] = PadroesRepository.contar()
                stats['padroes']['frequentes'] = PadroesRepository.contar_por_status('frequente')
                stats['padroes']['atrasados'] = PadroesRepository.contar_por_status('atrasado')
                stats['padroes']['faltantes'] = PadroesRepository.contar_por_status('faltante')

            if CombinacoesRepository is not None:
                stats['combinacoes']['total'] = CombinacoesRepository.contar()
                stats['combinacoes']['viaveis'] = CombinacoesRepository.contar(apenas_viaveis=True)
                stats['combinacoes']['nao_viaveis'] = CombinacoesRepository.contar(apenas_viaveis=False)

            if SorteiosReaisRepository is not None:
                stats['sorteios_analisados'] = SorteiosReaisRepository.contar()

        except Exception:
            pass

        return stats
