from services.analise_tubular_service import AnaliseTubularService
import random
import re
from datetime import datetime
import io
import requests


class GerarFechamentoTubularService:

    # Mapeamento de meses
    MESES_ABREV = {
        1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr',
        5: 'Mai', 6: 'Jun', 7: 'Jul', 8: 'Ago',
        9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
    }

    MESES_COMPLETOS = {
        'Janeiro': 1, 'Fevereiro': 2, 'Março': 3, 'Abril': 4,
        'Maio': 5, 'Junho': 6, 'Julho': 7, 'Agosto': 8,
        'Setembro': 9, 'Outubro': 10, 'Novembro': 11, 'Dezembro': 12
    }

    # ========================================================================
    # 🆕 FUNÇÕES DE PARSING PARA EXTRAIR VALORES DOS FORMATOS DO FRONTEND
    # ========================================================================

    @staticmethod
    def extrair_pares_de_par_impar(par_impar_str):
        """
        Extrai a quantidade de pares do formato do frontend.
        Formatos aceitos:
        - "3P + 4I" -> 3
        - "3P/4I" -> 3
        - "3P4I" -> 3
        - "3" -> 3
        """
        if not par_impar_str or not str(par_impar_str).strip():
            return None

        par_impar_str = str(par_impar_str).strip().upper()

        # Tenta extrair número antes de 'P'
        match = re.search(r'(\d+)\s*P', par_impar_str)
        if match:
            return int(match.group(1))

        # Tenta apenas número
        if par_impar_str.isdigit():
            return int(par_impar_str)

        return None

    @staticmethod
    def extrair_finais_iguais(finais_str):
        """
        Extrai a quantidade de finais iguais do formato do frontend.
        Formatos aceitos:
        - "2x final 1" -> (2, 1) -> quantidade=2, dígito final=1
        - "3x final 5" -> (3, 5)
        - "2" -> (2, None) -> apenas quantidade, qualquer final
        """
        if not finais_str or not str(finais_str).strip():
            return None, None

        finais_str = str(finais_str).strip().lower()

        # Tenta formato "Nx final D"
        match = re.search(r'(\d+)\s*x\s*final\s*(\d+)', finais_str)
        if match:
            return int(match.group(1)), int(match.group(2))

        # Tenta apenas número (quantidade de finais iguais, qualquer dígito)
        if finais_str.isdigit():
            return int(finais_str), None

        return None, None

    @staticmethod
    def extrair_digitos_unicos(digitos_str):
        """
        Extrai a quantidade de dígitos únicos do formato do frontend.
        Formatos aceitos:
        - "7 dígitos únicos" -> 7
        - "6 digitos unicos" -> 6
        - "7" -> 7
        """
        if not digitos_str or not str(digitos_str).strip():
            return None

        digitos_str = str(digitos_str).strip().lower()

        # Tenta formato "N dígitos únicos"
        match = re.search(r'^(\d+)\s*d[íi]gitos?\s*[úu]nicos?', digitos_str)
        if match:
            return int(match.group(1))

        # Tenta apenas número
        if digitos_str.replace(' ', '').isdigit():
            return int(digitos_str.strip())

        return None

    @staticmethod
    def extrair_tamanho_sequencia(sequencia_str):
        """
        Extrai o tamanho da sequência do formato do frontend.
        Formatos aceitos:
        - "Sequência de 2" -> 2
        - "Sequência de 3" -> 3
        - "Sequência de 4+" -> 4
        - "2" -> 2
        """
        if not sequencia_str or not str(sequencia_str).strip():
            return None

        sequencia_str = str(sequencia_str).strip()

        # Tenta formato "Sequência de N" ou "Sequência de N+"
        match = re.search(r'sequ[êe]ncia\s*de\s*(\d+)', sequencia_str, re.IGNORECASE)
        if match:
            return int(match.group(1))

        # Tenta apenas número
        if sequencia_str.isdigit():
            return int(sequencia_str)

        return None

    @staticmethod
    def extrair_padrao_inicial(padrao_str):
        """
        Converte o padrão inicial para o formato usado internamente.
        Formatos aceitos:
        - "0:2 | 1:3 | 2:2" -> lista de contagens por dígito
        - "0 0 1 1 1 2 2" -> conta ocorrências
        Retorna um dict com {digito: quantidade}
        """
        if not padrao_str or not str(padrao_str).strip():
            return None

        padrao_str = str(padrao_str).strip()

        # Formato "0:2 | 1:3 | 2:2"
        if ':' in padrao_str:
            resultado = {}
            partes = padrao_str.split('|')
            for parte in partes:
                match = re.search(r'(\d+)\s*:\s*(\d+)', parte.strip())
                if match:
                    digito = int(match.group(1))
                    quantidade = int(match.group(2))
                    resultado[digito] = quantidade
            return resultado if resultado else None

        # Formato "0 0 1 1 1 2 2" - conta ocorrências
        digitos = re.findall(r'\d', padrao_str)
        if digitos:
            resultado = {}
            for d in digitos:
                d_int = int(d)
                resultado[d_int] = resultado.get(d_int, 0) + 1
            return resultado

        return None

    # ========================================================================
    # MÉTODOS PRINCIPAIS
    # ========================================================================

    @staticmethod
    def obter_opcoes_para_fechamento():
        analise = AnaliseTubularService.obter_analise_completa()

        if 'erro' in analise:
            return {'erro': analise['erro']}

        # 🆕 Buscar dados das novas APIs
        soma_dezenas = GerarFechamentoTubularService.buscar_soma_dezenas()
        repeticoes_dezenas = GerarFechamentoTubularService.buscar_repeticoes_dezenas()
        padrao_inicial = GerarFechamentoTubularService.buscar_padrao_inicial()

        return {
            'sequencias': GerarFechamentoTubularService.extrair_top3(analise['sequencias']['padroes']),
            'finais': GerarFechamentoTubularService.extrair_top3(analise['finais'][:3]),
            'repeticoes': GerarFechamentoTubularService.formatar_repeticoes(analise['repeticoes']),
            'somas': GerarFechamentoTubularService.extrair_top3(analise['somas']['padroes'][:3]),
            'pares_impares': GerarFechamentoTubularService.extrair_top3(analise['pares_impares'][:3]),
            'padroes_iniciais_finais': GerarFechamentoTubularService.extrair_top3(analise['padroes_iniciais_finais'][:3]),
            'meses': GerarFechamentoTubularService.extrair_top3(analise['meses'][:3]),
            'digitos_unicos': GerarFechamentoTubularService.extrair_top3(analise['digitos_unicos'][:3]),
            # 🆕 Novos parâmetros
            'soma_dezenas': soma_dezenas,
            'repeticoes_dezenas': repeticoes_dezenas,
            'padrao_inicial': padrao_inicial,
            'recomendacao': GerarFechamentoTubularService.gerar_recomendacao(analise)
        }

    @staticmethod
    def extrair_top3(lista_padroes):
        resultado = []
        for padrao in lista_padroes[:3]:
            resultado.append({
                'descricao': padrao['descricao'],
                'frequencia': padrao['frequencia'],
                'percentual': padrao['percentual'],
                'dados': padrao
            })
        return resultado

    @staticmethod
    def formatar_repeticoes(repeticoes_data):
        return [{
            'descricao': f"Média de {repeticoes_data['media_repeticoes']} repetições",
            'frequencia': repeticoes_data['total'],
            'percentual': repeticoes_data['percentual'],
            'dados': repeticoes_data
        }]

    @staticmethod
    def gerar_recomendacao(analise):
        recomendacoes = []

        seq_top = analise['sequencias']['padroes'][0] if analise['sequencias']['padroes'] else None
        if seq_top:
            recomendacoes.append({
                'tipo': 'Sequências',
                'sugestao': f"Incluir {seq_top['descricao']} ({seq_top['percentual']}%)",
                'prioridade': 'ALTA' if seq_top['status'] == 'MAIS' else 'MÉDIA'
            })

        par_impar_top = analise['pares_impares'][0] if analise['pares_impares'] else None
        if par_impar_top:
            recomendacoes.append({
                'tipo': 'Par/Ímpar',
                'sugestao': f"Usar padrão {par_impar_top['descricao']} ({par_impar_top['percentual']}%)",
                'prioridade': 'ALTA' if par_impar_top['status'] == 'MAIS' else 'MÉDIA'
            })

        mes_top = analise['meses'][0] if analise['meses'] else None
        if mes_top:
            recomendacoes.append({
                'tipo': 'Mês da Sorte',
                'sugestao': f"Priorizar {mes_top['descricao']} ({mes_top['percentual']}%)",
                'prioridade': 'ALTA' if mes_top['status'] == 'MAIS' else 'MÉDIA'
            })

        return recomendacoes

    @staticmethod
    def gerar_jogos(parametros):
        """
        🆕 MODIFICADO: Agora retorna formato completo com filtros aplicados
        """
        quantidade = parametros.get('quantidade', 5)

        # Extrai filtros
        sequencia_escolhida = parametros.get('sequencia')
        par_impar_escolhido = parametros.get('par_impar')
        mes_escolhido = parametros.get('mes')
        digitos_unicos_escolhido = parametros.get('digitos_unicos')
        finais_iguais_escolhido = parametros.get('finais_iguais')

        # 🆕 Novos filtros
        numeros_excluir = parametros.get('excluir_numero', [])
        numeros_fixar = parametros.get('fixar_numero', [])
        faixa_min = parametros.get('faixa_min', 1)
        faixa_max = parametros.get('faixa_max', 31)

        # 🆕 Filtros dos 8 parâmetros
        padrao_inicial = parametros.get('padrao_inicial')
        soma_min = parametros.get('soma_min')
        soma_max = parametros.get('soma_max')
        repeticao_dezena = parametros.get('repeticao_dezena')

        jogos_gerados = []

        for i in range(quantidade):
            jogo = GerarFechamentoTubularService.gerar_jogo_unico(
                sequencia_escolhida,
                par_impar_escolhido,
                mes_escolhido,
                digitos_unicos_escolhido,
                finais_iguais_escolhido,
                numeros_excluir,
                numeros_fixar,
                faixa_min,
                faixa_max,
                padrao_inicial,
                soma_min,
                soma_max,
                repeticao_dezena
            )

            # Converte números para lista de strings para facilitar exportação
            numeros_str = [f"{n:02d}" for n in jogo['numeros']]

            jogos_gerados.append({
                'numero': i + 1,
                'numeros': jogo['numeros'],
                'numeros_str': numeros_str,  # 🆕 Versão formatada
                'mes_sorte': jogo['mes_sorte'],
                'mes_abrev': GerarFechamentoTubularService.MESES_ABREV[jogo['mes_sorte']],  # 🆕
                'analise': jogo['analise']
            })

        # 🆕 Retorna formato completo
        return {
            'ok': True,  # 🆕
            'total_jogos': quantidade,
            'total': len(jogos_gerados),  # 🆕
            'jogos': jogos_gerados,
            'apostas': [[f"{n:02d}" for n in jogo['numeros']] for jogo in jogos_gerados],  # 🆕 Formato array
            'filtros': {  # 🆕 Filtros aplicados
                'quantidade': quantidade,
                'sequencia': sequencia_escolhida,
                'par_impar': par_impar_escolhido,
                'mes': mes_escolhido,
                'digitos_unicos': digitos_unicos_escolhido,
                'finais_iguais': finais_iguais_escolhido,
                'excluir': numeros_excluir,
                'fixar': numeros_fixar,
                'faixa': [faixa_min, faixa_max]
            },
            'parametros_usados': parametros
        }

    @staticmethod
    def gerar_jogo_unico(sequencia, par_impar, mes, digitos_unicos, finais_iguais=None,
                         numeros_excluir=None, numeros_fixar=None, faixa_min=1, faixa_max=31,
                         padrao_inicial=None, soma_min=None, soma_max=None, repeticao_dezena=None):
        """
        🆕 MODIFICADO: Aplica TODOS os filtros rigorosamente (8 parâmetros)
        Usa as funções de parsing para extrair valores dos formatos do frontend
        """
        if numeros_excluir is None:
            numeros_excluir = []
        if numeros_fixar is None:
            numeros_fixar = []

        # 🆕 PARSING DOS VALORES DO FRONTEND
        pares_necessarios_total = GerarFechamentoTubularService.extrair_pares_de_par_impar(par_impar)
        qtd_finais_esperada, digito_final_esperado = GerarFechamentoTubularService.extrair_finais_iguais(finais_iguais)
        qtd_digitos_esperada = GerarFechamentoTubularService.extrair_digitos_unicos(digitos_unicos)
        tamanho_sequencia = GerarFechamentoTubularService.extrair_tamanho_sequencia(sequencia)
        padrao_inicial_dict = GerarFechamentoTubularService.extrair_padrao_inicial(padrao_inicial)

        # Parsing de repetição
        qtd_repeticoes_esperada = None
        if repeticao_dezena and str(repeticao_dezena).strip():
            rep_str = str(repeticao_dezena).strip()
            if rep_str.isdigit():
                qtd_repeticoes_esperada = int(rep_str)

        max_tentativas = 10000
        for tentativa in range(max_tentativas):
            numeros = []

            # 1️⃣ Adiciona números FIXOS primeiro
            numeros.extend(numeros_fixar)

            # 2️⃣ Calcula quantos números faltam
            numeros_faltantes = 7 - len(numeros)

            # 3️⃣ Cria pool de números disponíveis (respeitando FAIXA e EXCLUSÕES)
            pool_disponivel = [
                n for n in range(faixa_min, faixa_max + 1)
                if n not in numeros_excluir and n not in numeros
            ]

            # 4️⃣ Aplica filtro PAR/ÍMPAR (com parsing corrigido)
            if pares_necessarios_total is not None and numeros_faltantes > 0:
                impares_necessarios = 7 - pares_necessarios_total

                # Desconta os fixos
                pares_fixos = sum(1 for n in numeros if n % 2 == 0)
                impares_fixos = sum(1 for n in numeros if n % 2 != 0)

                pares_necessarios = pares_necessarios_total - pares_fixos
                impares_necessarios = impares_necessarios - impares_fixos

                pares_disponiveis = [n for n in pool_disponivel if n % 2 == 0]
                impares_disponiveis = [n for n in pool_disponivel if n % 2 != 0]

                if len(pares_disponiveis) >= pares_necessarios and len(impares_disponiveis) >= impares_necessarios:
                    if pares_necessarios > 0:
                        numeros.extend(random.sample(pares_disponiveis, pares_necessarios))
                    if impares_necessarios > 0:
                        impares_disponiveis_restantes = [n for n in impares_disponiveis if n not in numeros]
                        if len(impares_disponiveis_restantes) >= impares_necessarios:
                            numeros.extend(random.sample(impares_disponiveis_restantes, impares_necessarios))
                        else:
                            continue
                else:
                    continue  # Tenta novamente
            else:
                # Preenche com números aleatórios do pool
                if len(pool_disponivel) >= numeros_faltantes:
                    numeros.extend(random.sample(pool_disponivel, numeros_faltantes))
                else:
                    continue  # Não há números suficientes

            numeros.sort()

            # 5️⃣ Aplica filtro de SEQUÊNCIA (com parsing corrigido)
            if tamanho_sequencia is not None:
                numeros = GerarFechamentoTubularService.ajustar_para_sequencia(numeros, tamanho_sequencia)

            # 6️⃣ Valida FINAIS IGUAIS (com parsing corrigido)
            if qtd_finais_esperada is not None:
                if digito_final_esperado is not None:
                    # Precisa ter exatamente N números terminando com dígito específico
                    numeros_com_final = [n for n in numeros if n % 10 == digito_final_esperado]
                    if len(numeros_com_final) != qtd_finais_esperada:
                        continue  # Rejeita jogo
                else:
                    # Apenas verifica se tem N números com mesmo final (qualquer final)
                    qtd_finais_atual = GerarFechamentoTubularService.contar_finais_iguais(numeros)
                    if qtd_finais_atual < qtd_finais_esperada:
                        continue  # Rejeita jogo

            # 🆕 6b. Valida DÍGITOS ÚNICOS (com parsing corrigido)
            if qtd_digitos_esperada is not None:
                qtd_digitos_atual = GerarFechamentoTubularService.contar_digitos_unicos(numeros)
                if qtd_digitos_atual != qtd_digitos_esperada:
                    continue  # Rejeita jogo

            # 🆕 7️⃣ Valida SOMA (se especificado)
            soma_atual = sum(numeros)
            if soma_min is not None and soma_atual < soma_min:
                continue  # Soma muito baixa
            if soma_max is not None and soma_atual > soma_max:
                continue  # Soma muito alta

            # 🆕 8️⃣ Valida PADRÃO INICIAL (com parsing corrigido)
            if padrao_inicial_dict is not None:
                padrao_jogo = GerarFechamentoTubularService.calcular_padrao_digitos_dict(numeros)
                if padrao_jogo != padrao_inicial_dict:
                    continue  # Padrão não corresponde

            # 🆕 9️⃣ Valida REPETIÇÃO (se especificado)
            if qtd_repeticoes_esperada is not None:
                qtd_repeticoes_atual = GerarFechamentoTubularService.contar_repeticoes_ultimo_concurso(numeros)
                if qtd_repeticoes_atual != qtd_repeticoes_esperada:
                    continue  # Quantidade de repetições não corresponde

            # 🔟 Define MÊS DA SORTE
            mes_sorte = mes if mes else random.randint(1, 12)
            if isinstance(mes_sorte, str):
                mes_sorte = GerarFechamentoTubularService.obter_numero_mes(mes_sorte)

            # ✅ Jogo válido encontrado!
            return {
                'numeros': numeros,
                'mes_sorte': mes_sorte,
                'analise': {
                    'soma': sum(numeros),
                    'pares': len([n for n in numeros if n % 2 == 0]),
                    'impares': len([n for n in numeros if n % 2 != 0]),
                    'tem_sequencia': GerarFechamentoTubularService.verificar_sequencia(numeros)
                }
            }

        # ❌ Não conseguiu gerar após todas as tentativas
        raise Exception(f"Impossível gerar jogo válido com os filtros especificados após {max_tentativas} tentativas")

    @staticmethod
    def ajustar_para_sequencia(numeros, tamanho_seq):
        tentativas = 0
        while not GerarFechamentoTubularService.tem_sequencia_tamanho(numeros, tamanho_seq) and tentativas < 50:
            # Escolhe um índice aleatório para criar a sequência
            idx = random.randint(0, len(numeros) - tamanho_seq)
            base = numeros[idx]

            # Tenta criar sequência começando do número base
            nova_seq = list(range(base, base + tamanho_seq))

            # Verifica se todos os números da sequência são válidos (1-31)
            if all(1 <= n <= 31 for n in nova_seq):
                # Substitui os números necessários
                for i, novo_num in enumerate(nova_seq[1:], 1):
                    if idx + i < len(numeros):
                        numeros[idx + i] = novo_num

            numeros = list(set(numeros))  # Remove duplicatas
            while len(numeros) < 7:
                novo = random.randint(1, 31)
                if novo not in numeros:
                    numeros.append(novo)

            numeros.sort()
            tentativas += 1

        return numeros[:7]  # Garante exatamente 7 números

    @staticmethod
    def tem_sequencia_tamanho(numeros, tamanho):
        if len(numeros) < tamanho:
            return False
        contador = 1
        for i in range(1, len(numeros)):
            if numeros[i] == numeros[i-1] + 1:
                contador += 1
                if contador >= tamanho:
                    return True
            else:
                contador = 1
        return False

    @staticmethod
    def verificar_sequencia(numeros):
        for i in range(1, len(numeros)):
            if numeros[i] == numeros[i-1] + 1:
                return True
        return False

    @staticmethod
    def obter_numero_mes(nome_mes):
        return GerarFechamentoTubularService.MESES_COMPLETOS.get(nome_mes, random.randint(1, 12))

    # ========================================================================
    # 🆕 FUNÇÕES AUXILIARES PARA VALIDAÇÃO DOS 8 PARÂMETROS
    # ========================================================================

    @staticmethod
    def calcular_padrao_digitos(numeros):
        """
        Calcula o padrão de dígitos iniciais (0-9) dos números
        Exemplo: [1, 5, 14, 16, 22, 25, 29] -> "0:2 | 1:2 | 2:3"
        Retorna no formato compatível com o que vem da API
        """
        from collections import Counter

        # Extrai o primeiro dígito de cada número
        digitos = [int(str(n)[0]) for n in numeros]

        # Conta quantas vezes cada dígito aparece
        contador = Counter(digitos)

        # Cria o padrão no formato "0:2 | 1:2 | 2:3"
        partes = []
        for digito in range(10):  # 0 a 9
            qtd = contador.get(digito, 0)
            if qtd > 0:  # Só inclui dígitos que aparecem
                partes.append(f"{digito}:{qtd}")

        return " | ".join(partes)

    @staticmethod
    def calcular_padrao_digitos_dict(numeros):
        """
        Calcula o padrão de dígitos iniciais como dicionário
        Exemplo: [1, 5, 14, 16, 22, 25, 29] -> {0: 2, 1: 2, 2: 3}
        """
        from collections import Counter

        # Extrai o primeiro dígito de cada número
        digitos = [int(str(n)[0]) for n in numeros]

        # Conta quantas vezes cada dígito aparece
        return dict(Counter(digitos))

    @staticmethod
    def contar_finais_iguais(numeros):
        """
        Conta quantos números têm o mesmo dígito final
        Exemplo: [1, 11, 21, 5, 15] -> 3 (três números terminam em 1)
        Retorna o MÁXIMO de números com finais iguais
        """
        from collections import Counter

        # Extrai o dígito final de cada número
        finais = [n % 10 for n in numeros]

        # Conta a frequência de cada final
        contador = Counter(finais)

        # Retorna o maior grupo de finais iguais
        return max(contador.values()) if contador else 0

    @staticmethod
    def contar_digitos_unicos(numeros):
        """
        Conta quantos dígitos únicos (0-9) aparecem nos números
        Exemplo: [1, 5, 14, 16, 22, 25, 29] -> dígitos 0,1,2,4,5,6,9 = 7 dígitos únicos
        """
        digitos_presentes = set()

        for num in numeros:
            # Converte o número para string e adiciona cada dígito ao set
            for digito in str(num):
                digitos_presentes.add(int(digito))

        return len(digitos_presentes)

    @staticmethod
    def obter_ultimo_concurso():
        """
        Busca os números do último concurso realizado
        Retorna lista de números ou None se não encontrar
        """
        try:
            from models.sorteio import Sorteio
            ultimo = Sorteio.query.order_by(Sorteio.concurso.desc()).first()
            if ultimo:
                return sorted(ultimo.get_posicoes_lista())
        except Exception as e:
            print(f"Erro ao buscar último concurso: {e}")

        return None

    @staticmethod
    def contar_repeticoes_ultimo_concurso(numeros):
        """
        Conta quantas dezenas se repetem em relação ao último concurso
        Exemplo:
          - Último concurso: [1, 5, 10, 15, 20, 25, 30]
          - Jogo atual: [1, 5, 7, 14, 21, 28, 31]
          - Repetições: 2 (números 1 e 5)
        """
        ultimo_concurso = GerarFechamentoTubularService.obter_ultimo_concurso()

        if not ultimo_concurso:
            return 0

        # Conta quantos números do jogo atual estão no último concurso
        repeticoes = len(set(numeros) & set(ultimo_concurso))

        return repeticoes

    # ========================================================================
    # 🆕 MÉTODOS PARA BUSCAR DADOS DAS NOVAS APIs
    # ========================================================================

    @staticmethod
    def buscar_soma_dezenas():
        """
        Busca dados REAIS de soma das dezenas
        Service: AnaliseSomaDezenasService.analisar_somas()
        API de fallback: /api/analise/soma-dezenas
        """
        try:
            # Tenta usar o service diretamente (mais rápido)
            from services.analise_soma_dezenas_service import AnaliseSomaDezenasService
            dados = AnaliseSomaDezenasService.analisar_somas()

            if dados and 'estatisticas' in dados:
                stats = dados['estatisticas']
                return {
                    'minimo': stats.get('minimo', 0),
                    'maximo': stats.get('maximo', 0),
                    'media': round(stats.get('media', 0), 1)
                }
        except Exception as e:
            print(f"⚠️ Erro ao buscar soma_dezenas via Service: {e}")
            # Tenta via HTTP como fallback
            try:
                response = requests.get('http://localhost:5050/api/analise/soma-dezenas', timeout=3)
                if response.status_code == 200:
                    dados = response.json()
                    if dados and 'estatisticas' in dados:
                        stats = dados['estatisticas']
                        return {
                            'minimo': stats.get('minimo', 0),
                            'maximo': stats.get('maximo', 0),
                            'media': round(stats.get('media', 0), 1)
                        }
            except Exception as e2:
                print(f"⚠️ Erro ao buscar soma_dezenas via HTTP: {e2}")

        # Retorna valores padrão apenas se ambos falharem
        print("⚠️ Usando valores padrão para soma_dezenas")
        return {'minimo': 49, 'maximo': 172, 'media': 98.5}

    @staticmethod
    def buscar_repeticoes_dezenas():
        """
        Busca top 3 dezenas mais frequentes via AnaliseTubularService
        Usa service diretamente pois já está importado no início
        """
        try:
            analise = AnaliseTubularService.obter_analise_completa()

            if analise and 'dezenas_frequentes' in analise:
                top3 = analise['dezenas_frequentes'][:3]
                resultado = []
                for item in top3:
                    resultado.append({
                        'descricao': f"Dezena {item.get('numero', 0):02d}",
                        'frequencia': item.get('frequencia', 0),
                        'percentual': item.get('percentual', 0),
                        'numero': item.get('numero', 0)
                    })
                return resultado
        except Exception as e:
            print(f"Erro ao buscar repeticoes_dezenas: {e}")

        return [
            {'descricao': 'Dezena 01', 'frequencia': 0, 'percentual': 0, 'numero': 1},
            {'descricao': 'Dezena 02', 'frequencia': 0, 'percentual': 0, 'numero': 2},
            {'descricao': 'Dezena 03', 'frequencia': 0, 'percentual': 0, 'numero': 3}
        ]

    @staticmethod
    def buscar_padrao_inicial():
        """
        Busca top 3 padrões REAIS de dígitos iniciais
        Service: AnaliseDigitoPadraoInicialFinalService.analisar_padroes()
        API de fallback: /api/analise/digito-padrao-inicial-final
        """
        try:
            # Tenta usar o service diretamente (mais rápido)
            from services.analise_digito_padrao_inicial_final_service import AnaliseDigitoPadraoInicialFinalService
            analise = AnaliseDigitoPadraoInicialFinalService.analisar_padroes()

            if analise and 'top_padroes_iniciais' in analise:
                padroes = analise['top_padroes_iniciais'][:3]
                resultado = []
                for item in padroes:
                    resultado.append({
                        'descricao': item.get('padrao', ''),
                        'frequencia': item.get('frequencia', 0),
                        'percentual': round(item.get('porcentagem', 0), 1)
                    })
                return resultado
        except Exception as e:
            print(f"⚠️ Erro ao buscar padrao_inicial via Service: {e}")
            # Tenta via HTTP como fallback
            try:
                response = requests.get('http://localhost:5050/api/analise/digito-padrao-inicial-final', timeout=3)
                if response.status_code == 200:
                    analise = response.json()
                    if analise and 'top_padroes_iniciais' in analise:
                        padroes = analise['top_padroes_iniciais'][:3]
                        resultado = []
                        for item in padroes:
                            resultado.append({
                                'descricao': item.get('padrao', ''),
                                'frequencia': item.get('frequencia', 0),
                                'percentual': round(item.get('porcentagem', 0), 1)
                            })
                        return resultado
            except Exception as e2:
                print(f"⚠️ Erro ao buscar padrao_inicial via HTTP: {e2}")

        # Retorna valores padrão apenas se ambos falharem
        print("⚠️ Usando valores padrão para padrao_inicial")
        return [
            {'descricao': '0-1-2', 'frequencia': 0, 'percentual': 0},
            {'descricao': '1-2-3', 'frequencia': 0, 'percentual': 0},
            {'descricao': '0-0-1', 'frequencia': 0, 'percentual': 0}
        ]

    # ========================================================================
    # 🆕 FUNÇÕES DE EXPORTAÇÃO
    # ========================================================================

    @staticmethod
    def exportar_txt(jogos_data):
        """
        Exporta palpites em formato TXT simples
        Formato: 01 02 03 04 05 06 07 Fev
        """
        linhas = []

        for jogo in jogos_data['jogos']:
            numeros_str = ' '.join([f"{n:02d}" for n in jogo['numeros']])
            mes_abrev = GerarFechamentoTubularService.MESES_ABREV[jogo['mes_sorte']]
            linha = f"{numeros_str} {mes_abrev}"
            linhas.append(linha)

        conteudo_txt = '\n'.join(linhas)
        return conteudo_txt

    @staticmethod
    def exportar_xls(jogos_data):
        """
        Exporta palpites em formato XLS (Excel)
        Retorna bytes do arquivo Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise Exception("Biblioteca openpyxl não instalada. Execute: pip install openpyxl")

        # Cria workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Palpites Tubular"

        # Cabeçalhos
        headers = ['Numero1', 'Numero2', 'Numero3', 'Numero4', 'Numero5', 'Numero6', 'Numero7', 'Mês']

        # Estiliza cabeçalhos
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adiciona jogos
        for row_idx, jogo in enumerate(jogos_data['jogos'], start=2):
            # Números
            for col, numero in enumerate(jogo['numeros'], start=1):
                ws.cell(row=row_idx, column=col, value=numero)

            # Mês
            mes_abrev = GerarFechamentoTubularService.MESES_ABREV[jogo['mes_sorte']]
            ws.cell(row=row_idx, column=8, value=mes_abrev)

        # Ajusta largura das colunas
        for col in range(1, 9):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

        # Salva em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def exportar_html(jogos_data, parametros):
        """
        Exporta palpites em formato HTML completo
        Inclui filtros, jogos e análises
        """
        filtros = jogos_data.get('filtros', {})

        html = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Fechamento Tubular - {datetime.now().strftime('%d/%m/%Y %H:%M')}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        }}
        .container {{
            background: white;
            border-radius: 15px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
        }}
        h1 {{
            color: #667eea;
            text-align: center;
            margin-bottom: 10px;
        }}
        .subtitle {{
            text-align: center;
            color: #666;
            margin-bottom: 30px;
        }}
        .filtros {{
            background: #f8f9fa;
            border-left: 4px solid #667eea;
            padding: 15px;
            margin-bottom: 30px;
            border-radius: 5px;
        }}
        .filtros h2 {{
            color: #667eea;
            margin-top: 0;
            font-size: 18px;
        }}
        .filtro-item {{
            margin: 8px 0;
            color: #333;
        }}
        .filtro-label {{
            font-weight: bold;
            color: #667eea;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background: #667eea;
            color: white;
            padding: 12px;
            text-align: center;
            font-weight: 600;
        }}
        td {{
            padding: 10px;
            text-align: center;
            border-bottom: 1px solid #dee2e6;
        }}
        tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        tr:hover {{
            background: #e9ecef;
        }}
        .numero {{
            display: inline-block;
            width: 35px;
            height: 35px;
            line-height: 35px;
            background: #667eea;
            color: white;
            border-radius: 50%;
            margin: 0 3px;
            font-weight: bold;
        }}
        .mes {{
            background: #28a745;
            color: white;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
        }}
        .footer {{
            text-align: center;
            color: #666;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 1px solid #dee2e6;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        .stat-box {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }}
        .stat-number {{
            font-size: 32px;
            font-weight: bold;
        }}
        .stat-label {{
            font-size: 14px;
            opacity: 0.9;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎲 Fechamento Tubular - Dia de Sorte</h1>
        <div class="subtitle">Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}</div>

        <div class="stats">
            <div class="stat-box">
                <div class="stat-number">{jogos_data['total']}</div>
                <div class="stat-label">Jogos Gerados</div>
            </div>
            <div class="stat-box">
                <div class="stat-number">7</div>
                <div class="stat-label">Números por Jogo</div>
            </div>
            <div class="stat-box">
                <div class="stat-number">{sum([jogo['analise']['soma'] for jogo in jogos_data['jogos']]) // jogos_data['total']}</div>
                <div class="stat-label">Soma Média</div>
            </div>
        </div>

        <div class="filtros">
            <h2>📋 Filtros Aplicados</h2>
            <div class="filtro-item"><span class="filtro-label">Quantidade:</span> {filtros.get('quantidade', 'Não especificado')}</div>
            <div class="filtro-item"><span class="filtro-label">Sequência:</span> {filtros.get('sequencia') or 'Não especificado'}</div>
            <div class="filtro-item"><span class="filtro-label">Par/Ímpar:</span> {filtros.get('par_impar') or 'Não especificado'}</div>
            <div class="filtro-item"><span class="filtro-label">Mês da Sorte:</span> {filtros.get('mes') or 'Aleatório'}</div>
            <div class="filtro-item"><span class="filtro-label">Dígitos Únicos:</span> {filtros.get('digitos_unicos') or 'Não especificado'}</div>
            {f'<div class="filtro-item"><span class="filtro-label">Números Excluídos:</span> {", ".join(map(str, filtros.get("excluir", []))) or "Nenhum"}</div>' if filtros.get('excluir') else ''}
            {f'<div class="filtro-item"><span class="filtro-label">Números Fixos:</span> {", ".join(map(str, filtros.get("fixar", []))) or "Nenhum"}</div>' if filtros.get('fixar') else ''}
            <div class="filtro-item"><span class="filtro-label">Faixa:</span> {filtros.get('faixa', [1, 31])[0]} a {filtros.get('faixa', [1, 31])[1]}</div>
        </div>

        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Palpite</th>
                    <th>Mês</th>
                    <th>Soma</th>
                    <th>Pares</th>
                    <th>Ímpares</th>
                    <th>Seq.</th>
                </tr>
            </thead>
            <tbody>
"""

        # Adiciona jogos
        for jogo in jogos_data['jogos']:
            numeros_html = ''.join([f'<span class="numero">{n:02d}</span>' for n in jogo['numeros']])
            mes_abrev = GerarFechamentoTubularService.MESES_ABREV[jogo['mes_sorte']]

            html += f"""
                <tr>
                    <td>{jogo['numero']}</td>
                    <td>{numeros_html}</td>
                    <td><span class="mes">{mes_abrev}</span></td>
                    <td>{jogo['analise']['soma']}</td>
                    <td>{jogo['analise']['pares']}</td>
                    <td>{jogo['analise']['impares']}</td>
                    <td>{'✅' if jogo['analise']['tem_sequencia'] else '❌'}</td>
                </tr>
"""

        html += """
            </tbody>
        </table>

        <div class="footer">
            <p><strong>Sistema de Análise - Dia de Sorte</strong></p>
            <p>Este relatório foi gerado automaticamente pelo sistema de fechamento tubular.</p>
        </div>
    </div>
</body>
</html>
"""

        return html
