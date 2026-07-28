# Sistema: Gerador de Apostas por Padrões - Dia de Sorte
# Desenvolvido para: Márcio Fernando Maia
# Rotas da API para geração completa de apostas por padrões

from flask import Blueprint, jsonify, request, render_template, Response
from services.gerador_padroes_completo_service import GeradorPadroesCompletoService
import io

gerador_padroes_bp = Blueprint('gerador_padroes', __name__)


# =========================================================================
# PÁGINA PRINCIPAL
# =========================================================================

@gerador_padroes_bp.route('/gerador-padroes')
def pagina_gerador_padroes():
    """
    Página principal do gerador de apostas por padrões
    """
    return render_template('gerador_padroes_completo.html')


# =========================================================================
# API: LISTAGEM DE PADRÕES
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/listar', methods=['GET'])
def listar_padroes():
    """
    Lista todos os padrões viáveis com estatísticas.

    Query params:
        status: 'frequente', 'atrasado', 'faltante' (opcional)
        min_jogos: Mínimo de jogos possíveis (opcional)
        max_jogos: Máximo de jogos possíveis (opcional)
        ordenar: 'jogos_possiveis', 'frequencia', 'atraso' (default: jogos_possiveis)

    Returns:
        JSON com lista de padrões e estatísticas
    """
    try:
        # Parâmetros de filtro
        status = request.args.get('status', None)
        min_jogos = request.args.get('min_jogos', None)
        max_jogos = request.args.get('max_jogos', None)
        ordenar = request.args.get('ordenar', 'jogos_possiveis')

        # Converter para int se fornecidos
        if min_jogos:
            min_jogos = int(min_jogos)
        if max_jogos:
            max_jogos = int(max_jogos)

        # Se há filtros específicos, usar filtrar_padroes
        if status or min_jogos or max_jogos:
            padroes = GeradorPadroesCompletoService.filtrar_padroes(
                status=status,
                min_jogos=min_jogos,
                max_jogos=max_jogos,
                ordenar_por=ordenar
            )

            total_jogos = sum(p['jogos_possiveis'] for p in padroes)

            # Contar por status
            frequentes = sum(1 for p in padroes if p['status'] == 'frequente')
            atrasados = sum(1 for p in padroes if p['status'] == 'atrasado')
            faltantes = sum(1 for p in padroes if p['status'] == 'faltante')

            ultimo_resultado = GeradorPadroesCompletoService.obter_ultimo_resultado_padrao()
            padrao_ult = ultimo_resultado['padrao'] if ultimo_resultado else None
            for p in padroes:
                p['eh_padrao_ultimo_concurso'] = (p.get('padrao') == padrao_ult)

            total_sorteios = ultimo_resultado['concurso'] if ultimo_resultado else 0
            top_frequencia = GeradorPadroesCompletoService.obter_top_frequencia(
                padroes, limite=3, total_sorteios=total_sorteios
            )

            return jsonify({
                'sucesso': True,
                'padroes': padroes,
                'total_padroes': len(padroes),
                'total_jogos_possiveis': total_jogos,
                'filtros_aplicados': True,
                'ultimo_resultado': ultimo_resultado,
                'total_sorteios_analisados': total_sorteios,
                'top_frequencia': top_frequencia,
                'padrao_mais_frequente': top_frequencia[0] if top_frequencia else None,
                'estatisticas': {
                    'frequentes': frequentes,
                    'atrasados': atrasados,
                    'faltantes': faltantes
                }
            })

        # Sem filtros, retornar todos
        dados = GeradorPadroesCompletoService.listar_padroes_com_historico()

        return jsonify({
            'sucesso': True,
            **dados
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/top-frequencia-elite', methods=['GET'])
def top_frequencia_elite():
    """
    Top padrões posicionais mais frequentes + concursos históricos por padrão.
    Usado pelo Simulador Elite (modo 3 linhas padrão + 7 relações dígitos/soma).
    """
    try:
        limite = int(request.args.get('limite', 3))
        dados = GeradorPadroesCompletoService.obter_top_frequencia_elite(limite=limite)
        return jsonify({
            'sucesso': True,
            **dados,
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e),
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/estatisticas', methods=['GET'])
def estatisticas_padroes():
    """
    Retorna estatísticas gerais sobre todos os padrões.
    """
    try:
        stats = GeradorPadroesCompletoService.obter_estatisticas_gerais()

        return jsonify({
            'sucesso': True,
            **stats
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/buscar', methods=['GET'])
def buscar_padrao():
    """
    Busca informações de um padrão específico.

    Query params:
        padrao: String do padrão (ex: "0 0 0 0 0 2 3")

    Returns:
        JSON com informações completas do padrão
    """
    padrao_str = request.args.get('padrao', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        resultado = GeradorPadroesCompletoService.buscar_padrao(padrao_str)

        if resultado:
            return jsonify({
                'sucesso': True,
                'padrao': resultado
            })
        else:
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável ou inválido'
            }), 404

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: GERAÇÃO DE APOSTAS
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/gerar', methods=['GET'])
def gerar_apostas():
    """
    Gera apostas para um padrão específico.

    Query params:
        padrao: String do padrão (obrigatório)
        limite: Limite de jogos a retornar (opcional, default: todos)
        pagina: Número da página (opcional, para paginação)
        por_pagina: Jogos por página (opcional, default: 100)

    Returns:
        JSON com jogos gerados
    """
    padrao_str = request.args.get('padrao', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        limite = request.args.get('limite', None)
        pagina = request.args.get('pagina', None)
        por_pagina = int(request.args.get('por_pagina', 100))

        # Verificar se padrão é viável
        total_possivel = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
        if total_possivel == 0:
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável'
            }), 400

        # Se paginação solicitada
        if pagina:
            pagina = int(pagina)
            resultado = GeradorPadroesCompletoService.gerar_combinacoes_paginado(
                padrao_str, pagina, por_pagina
            )

            return jsonify({
                'sucesso': True,
                **resultado
            })

        # Geração completa (com limite opcional)
        resultado = GeradorPadroesCompletoService.gerar_todas_combinacoes(padrao_str)

        # Aplicar limite se especificado
        if limite:
            limite = int(limite)
            if limite == 0:
                limite = None
            else:
                resultado['jogos'] = resultado['jogos'][:limite]
                resultado['total'] = len(resultado['jogos'])

        return jsonify({
            'sucesso': True,
            **resultado
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/gerar-com-desdobramento', methods=['GET'])
def gerar_apostas_com_desdobramento():
    """
    Gera apostas para um padrão garantindo pelo menos 2 dezenas do último sorteio.
    
    Query params:
        padrao: String do padrão (obrigatório)
        limite: Limite de jogos a retornar (opcional, default: todos)
        sorteio_base: Lista de números do último sorteio separados por vírgula (obrigatório)
        
    Retorna:
        JSON com jogos gerados que contêm mínimo 2 dezenas do sorteio base
    """
    padrao_str = request.args.get('padrao', '')
    sorteio_base_str = request.args.get('sorteio_base', '')
    
    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400
        
    if not sorteio_base_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "sorteio_base" é obrigatório'
        }), 400
    
    try:
        # Parser sorteio_base
        sorteio_base = [int(n.strip()) for n in sorteio_base_str.split(',') if n.strip()]
        sorteio_set = set(sorteio_base)
        
        if len(sorteio_base) != 7:
            return jsonify({
                'sucesso': False,
                'erro': 'Sorteio base deve conter exatamente 7 números'
            }), 400
        
        # Verificar viabilidade do padrão
        total_possivel = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)
        if total_possivel == 0:
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável'
            }), 400
        
        # Gerar todas as combinações
        resultado = GeradorPadroesCompletoService.gerar_todas_combinacoes(padrao_str)
        jogos_originais = resultado.get('jogos', [])
        
        # Filtrar: manter apenas apostas com >= 2 números do sorteio base
        jogos_filtrados = []
        for jogo in jogos_originais:
            jogo_set = set(jogo)
            intersecao = len(jogo_set & sorteio_set)
            if intersecao >= 2:
                jogos_filtrados.append(jogo)
        
        limite = request.args.get('limite', None)
        if limite:
            limite = int(limite)
            if limite > 0:
                jogos_filtrados = jogos_filtrados[:limite]
        
        # Retornar resultado
        tempo_geracao = resultado.get('tempo_geracao', 0)
        
        return jsonify({
            'sucesso': True,
            'jogos': jogos_filtrados,
            'total': len(jogos_filtrados),
            'total_original': len(jogos_originais),
            'filtrados_por_desdobramento': True,
            'sorteio_base': sorteio_base,
            'tempo_geracao': tempo_geracao,
            'taxa_retencao': round(len(jogos_filtrados) / len(jogos_originais) * 100, 1) if jogos_originais else 0
        })
    
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/gerar-multiplo', methods=['GET'])
def gerar_apostas_multiplas():
    """
    Gera apostas para múltiplos padrões simultaneamente.
    🆕 VERSÃO ATUALIZADA: Usa baixas/médias/altas e respeita dezenas selecionadas
    🆕 VERSÃO 2: Distribui quantidade IGUALMENTE entre padrões
    🆕 VERSÃO 3: SUPORTE A DESDOBRAMENTO - Filtra por ≥2 números do último sorteio

    Query params:
        padroes: Padrões separados por vírgula (obrigatório)
        quantidade: Quantidade de apostas a gerar (opcional, default: 10)
        dezenas: Dezenas a usar (separadas por vírgula, ou 'todas') (opcional, default: todas)
        sorteio_base: Lista de números do último sorteio (opcional, para desdobramento)

    Returns:
        JSON com jogos gerados combinando todos os padrões EQUILIBRADAMENTE
    """
    padroes_str = request.args.get('padroes', '')
    quantidade = request.args.get('quantidade', '10')
    dezenas_str = request.args.get('dezenas', 'todas')
    sorteio_base_str = request.args.get('sorteio_base', '')  # 🆕 Parâmetro para desdobramento

    if not padroes_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padroes" é obrigatório'
        }), 400

    try:
        # Converter quantidade para int
        try:
            quantidade = int(quantidade)
            if quantidade < 1:
                quantidade = 10
        except ValueError:
            quantidade = 10

        # Processar padrões
        padroes_lista = [p.strip() for p in padroes_str.split(',')]
        padroes_lista = [p for p in padroes_lista if p]

        if not padroes_lista:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum padrão válido fornecido'
            }), 400

        # 🆕 Processar sorteio base para desdobramento
        sorteio_base = None
        usar_desdobramento = False
        if sorteio_base_str:
            try:
                sorteio_base = [int(n.strip()) for n in sorteio_base_str.split(',') if n.strip()]
                if len(sorteio_base) == 7:
                    usar_desdobramento = True
            except ValueError:
                sorteio_base = None

        # 🆕 Processar dezenas selecionadas
        dezenas_selecionadas = None
        usar_todas = dezenas_str.lower() == 'todas'
        
        if not usar_todas:
            try:
                dezenas_selecionadas = [int(d.strip()) for d in dezenas_str.split(',') if d.strip()]
                if not dezenas_selecionadas:
                    dezenas_selecionadas = list(range(1, 32))  # Se vazio, usar todas
                else:
                    # Validar dezenas (1-31)
                    dezenas_selecionadas = [d for d in dezenas_selecionadas if 1 <= d <= 31]
            except ValueError:
                dezenas_selecionadas = list(range(1, 32))
        else:
            dezenas_selecionadas = list(range(1, 32))

        # 🆕 DISTRIBUIÇÃO EQUILIBRADA
        # Dividir a quantidade igualmente entre os padrões
        qtd_padroes = len(padroes_lista)
        qtd_por_padrao = quantidade // qtd_padroes
        resto = quantidade % qtd_padroes

        # Gerar apostas para cada padrão de forma equilibrada
        jogos_por_padrao = {}
        tempo_geracao = 0
        padroes_processados = 0
        erros = []

        for idx, padrao in enumerate(padroes_lista):
            try:
                # Calcular quantidade para este padrão
                # Os primeiros 'resto' padrões recebem +1 aposta
                qtd_este_padrao = qtd_por_padrao + (1 if idx < resto else 0)

                # 🆕 LOG: Debug do padrão sendo processado
                print(f"\n🔍 Processando padrão '{padrao}'")
                print(f"   Dezenas selecionadas: {dezenas_selecionadas}")
                print(f"   Quantidade para este padrão: {qtd_este_padrao}")

                # 🆕 Usar nova função que respeita baixas/médias/altas
                resultado = GeradorPadroesCompletoService.gerar_com_dezenas_selecionadas(
                    padrao, dezenas_selecionadas
                )
                
                if 'erro' in resultado:
                    erro_msg = f"Padrão '{padrao}': {resultado['erro']}"
                    print(f"   ❌ ERRO: {erro_msg}")
                    erros.append(erro_msg)
                    continue
                
                print(f"   ✅ Gerou {len(resultado.get('jogos', []))} jogos")

                jogos = resultado.get('jogos', [])

                # Garantir apenas jogos com 7 dezenas (defensivo)
                jogos_validos = [j for j in jogos if len(j) == 7]
                jogos_invalidos = len(jogos) - len(jogos_validos)
                if jogos_invalidos and not jogos_validos:
                    erros.append(f"Padrão '{padrao}': nenhuma combinação válida de 7 dezenas (descartadas {jogos_invalidos})")
                    continue
                if jogos_invalidos:
                    erros.append(f"Padrão '{padrao}': descartadas {jogos_invalidos} combinações com tamanho diferente de 7")
                jogos = jogos_validos
                tempo_geracao += resultado.get('tempo_geracao', 0)

                # 🆕 FILTRO DE DESDOBRAMENTO: Manter apenas jogos com >= 2 números do sorteio base
                if usar_desdobramento and sorteio_base:
                    sorteio_set = set(sorteio_base)
                    jogos_filtrados = []
                    for jogo in jogos:
                        jogo_set = set(jogo)
                        intersecao = len(jogo_set & sorteio_set)
                        if intersecao >= 2:
                            jogos_filtrados.append(jogo)
                    
                    total_antes = len(jogos)
                    jogos = jogos_filtrados
                    total_depois = len(jogos)
                    
                    if total_depois == 0:
                        erros.append(f"Padrão '{padrao}': nenhum jogo com ≥2 números do desdobramento (de {total_antes} jogos)")
                        continue
                    else:
                        erros.append(f"Padrão '{padrao}': {total_depois} de {total_antes} jogos mantidos com desdobramento")

                padroes_processados += 1

                # 🆕 Embaralhar e pegar exatamente qtd_este_padrao
                import random
                random.shuffle(jogos)
                jogos_selecionados = jogos[:qtd_este_padrao]
                
                jogos_por_padrao[padrao] = jogos_selecionados

            except Exception as e:
                erros.append(f"Padrão '{padrao}': {str(e)}")
                continue

        if padroes_processados == 0:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum padrão válido pôde ser processado',
                'detalhes': erros
            }), 400

        # 🆕 Combinar jogos de forma alternada para melhor distribuição visual
        todos_jogos = []
        max_jogos = max(len(jogos) for jogos in jogos_por_padrao.values())
        
        for i in range(max_jogos):
            for padrao in padroes_lista:
                if i < len(jogos_por_padrao[padrao]):
                    todos_jogos.append(jogos_por_padrao[padrao][i])

        # Remover duplicatas mantendo ordem
        jogos_unicos = []
        vistos = set()
        for jogo in todos_jogos:
            jogo_tuple = tuple(sorted(jogo))
            if jogo_tuple not in vistos:
                vistos.add(jogo_tuple)
                jogos_unicos.append(jogo)

        # 🆕 Preparar resposta
        resposta = {
            'sucesso': True,
            'jogos': jogos_unicos[:quantidade],  # Garantir exatamente 'quantidade'
            'total': len(jogos_unicos[:quantidade]),
            'padroes_processados': padroes_processados,
            'distribuicao': {
                padrao: len(jogos_por_padrao.get(padrao, []))
                for padrao in padroes_lista
            },
            'tempo_geracao': round(tempo_geracao, 2),
            'erros': erros if erros else None,
            'info': f'Gerado usando {len(dezenas_selecionadas)} dezenas selecionadas - {qtd_por_padrao} apostas por padrão (mais {resto} para os primeiros)',
            'aviso_filtragem': 'Combinações com menos de 7 dezenas são descartadas automaticamente' if erros else None,
            'debug_dezenas': dezenas_selecionadas,  # 🆕 DEBUG
            'debug_sorteio_base': sorteio_base,     # 🆕 DEBUG
            'debug_usar_desdobramento': usar_desdobramento  # 🆕 DEBUG
        }

        # 🆕 Adicionar informação de desdobramento se aplicável
        if usar_desdobramento:
            resposta['filtrados_por_desdobramento'] = True
            resposta['sorteio_base'] = sorteio_base
            resposta['msg_desdobramento'] = f'Desdobramento ativo: apenas jogos com ≥2 números do último concurso ({", ".join(map(str, sorteio_base))})'

        return jsonify(resposta)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/calcular', methods=['GET'])
def calcular_jogos():
    """
    Calcula quantos jogos são possíveis para um padrão.

    Query params:
        padrao: String do padrão

    Returns:
        JSON com quantidade de jogos possíveis
    """
    padrao_str = request.args.get('padrao', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        jogos = GeradorPadroesCompletoService.calcular_jogos_possiveis(padrao_str)

        return jsonify({
            'sucesso': True,
            'padrao': padrao_str,
            'jogos_possiveis': jogos,
            'viavel': jogos > 0
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: EXPORTAÇÃO
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/exportar-txt', methods=['GET'])
def exportar_txt():
    """
    Exporta todas as apostas de um padrão em formato TXT.

    Query params:
        padrao: String do padrão (obrigatório)
        mes: Abreviação do mês (opcional, ex: "Nov")

    Returns:
        Arquivo TXT para download
    """
    padrao_str = request.args.get('padrao', '')
    mes_abrev = request.args.get('mes', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        # Verificar viabilidade
        if not GeradorPadroesCompletoService.verificar_padrao_viavel(padrao_str):
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável'
            }), 400

        # Gerar conteúdo TXT
        conteudo = GeradorPadroesCompletoService.exportar_para_txt(padrao_str, mes_abrev)

        # Criar response como arquivo
        padrao_nome = padrao_str.replace(' ', '')

        return Response(
            conteudo,
            mimetype='text/plain',
            headers={
                'Content-Disposition': f'attachment; filename=padrao_{padrao_nome}.txt',
                'Content-Type': 'text/plain; charset=utf-8'
            }
        )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/exportar-xlsx', methods=['GET'])
def exportar_xlsx():
    """
    Exporta todas as apostas de um padrão em formato XLSX.

    Query params:
        padrao: String do padrão (obrigatório)
        mes: Nome do mês (opcional, ex: "Novembro")

    Returns:
        Arquivo XLSX para download
    """
    padrao_str = request.args.get('padrao', '')
    mes_nome = request.args.get('mes', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        # Verificar viabilidade
        if not GeradorPadroesCompletoService.verificar_padrao_viavel(padrao_str):
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável'
            }), 400

        # Gerar dados
        dados = GeradorPadroesCompletoService.exportar_para_xlsx_data(padrao_str, mes_nome)

        # Criar XLSX usando openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Apostas"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="D4B31A", end_color="D4B31A", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ['Jogo', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'Padrão']
            if mes_nome:
                headers.append('Mês')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Dados
            for row_idx, linha in enumerate(dados, 2):
                ws.cell(row=row_idx, column=1, value=linha['Jogo']).border = thin_border
                ws.cell(row=row_idx, column=2, value=linha['D1']).border = thin_border
                ws.cell(row=row_idx, column=3, value=linha['D2']).border = thin_border
                ws.cell(row=row_idx, column=4, value=linha['D3']).border = thin_border
                ws.cell(row=row_idx, column=5, value=linha['D4']).border = thin_border
                ws.cell(row=row_idx, column=6, value=linha['D5']).border = thin_border
                ws.cell(row=row_idx, column=7, value=linha['D6']).border = thin_border
                ws.cell(row=row_idx, column=8, value=linha['D7']).border = thin_border
                ws.cell(row=row_idx, column=9, value=linha['Padrão']).border = thin_border

                if mes_nome and 'Mês' in linha:
                    ws.cell(row=row_idx, column=10, value=linha['Mês']).border = thin_border

            # Ajustar larguras das colunas
            ws.column_dimensions['A'].width = 10
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 6
            ws.column_dimensions['I'].width = 20
            if mes_nome:
                ws.column_dimensions['J'].width = 12

            # Salvar em buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            padrao_nome = padrao_str.replace(' ', '')

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=padrao_{padrao_nome}.xlsx',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            )

        except ImportError:
            # Fallback: exportar como CSV se openpyxl não estiver disponível
            import csv

            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=list(dados[0].keys()) if dados else [])
            writer.writeheader()
            writer.writerows(dados)

            padrao_nome = padrao_str.replace(' ', '')

            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=padrao_{padrao_nome}.csv',
                    'Content-Type': 'text/csv; charset=utf-8'
                }
            )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: INFORMAÇÕES
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/info', methods=['GET'])
def info_gerador():
    """
    Retorna informações sobre o módulo de geração.
    """
    try:
        dados = GeradorPadroesCompletoService.listar_padroes_com_historico()

        return jsonify({
            'sucesso': True,
            'modulo': 'Gerador de Apostas por Padrões',
            'versao': '1.0',
            'total_padroes_viaveis': dados['total_padroes'],
            'total_jogos_possiveis': dados['total_jogos_possiveis'],
            'faixas': {
                0: {'range': '01-09', 'quantidade': 9},
                1: {'range': '10-19', 'quantidade': 10},
                2: {'range': '20-29', 'quantidade': 10},
                3: {'range': '30-31', 'quantidade': 2}
            },
            'formatos_exportacao': ['TXT', 'XLSX'],
            'ultimo_concurso': dados['ultimo_concurso']
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: PERSISTÊNCIA NO BANCO DE DADOS
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/salvar-padrao', methods=['POST'])
def salvar_padrao():
    """
    Salva um padrão e suas combinações no banco de dados.

    Body JSON:
        padrao: String do padrão (obrigatório)
        limite: Limite de combinações a salvar (opcional)

    Returns:
        JSON com resultado da operação
    """
    try:
        dados = request.get_json() or {}
        padrao_str = dados.get('padrao', '')

        if not padrao_str:
            return jsonify({
                'sucesso': False,
                'erro': 'Parâmetro "padrao" é obrigatório'
            }), 400

        limite = dados.get('limite', None)

        # Salvar padrão e combinações
        resultado = GeradorPadroesCompletoService.salvar_combinacoes_no_banco(
            padrao_str, limite=limite
        )

        if resultado.get('erro'):
            return jsonify({
                'sucesso': False,
                'erro': resultado['erro']
            }), 500

        return jsonify({
            'sucesso': True,
            'padrao': padrao_str,
            'combinacoes_salvas': resultado['salvas'],
            'combinacoes_existentes': resultado['existentes']
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/combinacoes-banco', methods=['GET'])
def obter_combinacoes_banco():
    """
    Obtém combinações do banco de dados com paginação.

    Query params:
        padrao: String do padrão (obrigatório)
        pagina: Número da página (default: 1)
        por_pagina: Itens por página (default: 100)
        apenas_viaveis: true/false (opcional)

    Returns:
        JSON com combinações e metadados de paginação
    """
    padrao_str = request.args.get('padrao', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 100))

        apenas_viaveis = request.args.get('apenas_viaveis', None)
        if apenas_viaveis == 'true':
            apenas_viaveis = True
        elif apenas_viaveis == 'false':
            apenas_viaveis = False
        else:
            apenas_viaveis = None

        resultado = GeradorPadroesCompletoService.obter_combinacoes_do_banco(
            padrao_str,
            pagina=pagina,
            por_pagina=por_pagina,
            apenas_viaveis=apenas_viaveis
        )

        return jsonify({
            'sucesso': True,
            **resultado
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/gerar-com-viabilidade', methods=['GET'])
def gerar_com_viabilidade():
    """
    Gera combinações com classificação de viabilidade.

    Query params:
        padrao: String do padrão (obrigatório)
        limite: Limite de combinações (opcional)

    Returns:
        JSON com combinações e informações de viabilidade
    """
    padrao_str = request.args.get('padrao', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        limite = request.args.get('limite', None)
        if limite:
            limite = int(limite)

        combinacoes = GeradorPadroesCompletoService.gerar_combinacoes_com_viabilidade(
            padrao_str, limite=limite
        )

        # Estatísticas de viabilidade
        total = len(combinacoes)
        viaveis = sum(1 for c in combinacoes if c['viavel'])
        nao_viaveis = total - viaveis

        return jsonify({
            'sucesso': True,
            'padrao': padrao_str,
            'total': total,
            'viaveis': viaveis,
            'nao_viaveis': nao_viaveis,
            'percentual_viaveis': round(viaveis / total * 100, 2) if total > 0 else 0,
            'combinacoes': combinacoes
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/estatisticas-banco', methods=['GET'])
def estatisticas_banco():
    """
    Retorna estatísticas do banco de dados.

    Returns:
        JSON com estatísticas de padrões, combinações e sorteios
    """
    try:
        stats = GeradorPadroesCompletoService.obter_estatisticas_banco()

        return jsonify({
            'sucesso': True,
            **stats
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: BUSCA INTELIGENTE PÓS-SORTEIO
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/analisar-sorteio', methods=['POST'])
def analisar_sorteio():
    """
    Analisa um resultado de sorteio real.

    Body JSON:
        dezenas: Lista de 7 dezenas sorteadas (obrigatório)
        concurso: Número do concurso (opcional)
        mes_sorte: Mês da sorte 1-12 (opcional)
        salvar: Se deve salvar no banco (default: true)

    Returns:
        JSON com análise completa do sorteio
    """
    try:
        dados = request.get_json() or {}
        dezenas = dados.get('dezenas', [])

        if not dezenas or len(dezenas) != 7:
            return jsonify({
                'sucesso': False,
                'erro': 'São necessárias exatamente 7 dezenas'
            }), 400

        concurso = dados.get('concurso', None)
        mes_sorte = dados.get('mes_sorte', None)
        salvar = dados.get('salvar', True)

        resultado = GeradorPadroesCompletoService.analisar_sorteio_real(
            dezenas,
            concurso=concurso,
            mes_sorte=mes_sorte,
            salvar=salvar
        )

        if 'erro' in resultado:
            return jsonify({
                'sucesso': False,
                'erro': resultado['erro']
            }), 400

        return jsonify({
            'sucesso': True,
            **resultado
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/historico-sorteios', methods=['GET'])
def historico_sorteios():
    """
    Retorna histórico de sorteios reais analisados.

    Query params:
        limite: Quantidade de sorteios (default: 50)

    Returns:
        JSON com lista de sorteios analisados
    """
    try:
        limite = int(request.args.get('limite', 50))

        sorteios = GeradorPadroesCompletoService.obter_historico_sorteios_reais(limite=limite)

        return jsonify({
            'sucesso': True,
            'total': len(sorteios),
            'sorteios': sorteios
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@gerador_padroes_bp.route('/api/gerador-padroes/avaliar-viabilidade', methods=['POST'])
def avaliar_viabilidade():
    """
    Avalia a viabilidade estatística de uma combinação.

    Body JSON:
        dezenas: Lista de 7 dezenas (obrigatório)

    Returns:
        JSON com avaliação de viabilidade
    """
    try:
        dados = request.get_json() or {}
        dezenas = dados.get('dezenas', [])

        if not dezenas or len(dezenas) != 7:
            return jsonify({
                'sucesso': False,
                'erro': 'São necessárias exatamente 7 dezenas'
            }), 400

        viavel, motivo = GeradorPadroesCompletoService.avaliar_viabilidade_combinacao(dezenas)

        return jsonify({
            'sucesso': True,
            'dezenas': sorted(dezenas),
            'viavel': viavel,
            'motivo': motivo,
            'recomendacao': 'Jogar' if viavel else 'Evitar'
        })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: EXPORTAÇÃO COM VIABILIDADE
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/exportar-xlsx-viabilidade', methods=['GET'])
def exportar_xlsx_viabilidade():
    """
    Exporta apostas em XLSX com informações de viabilidade.

    Query params:
        padrao: String do padrão (obrigatório)
        mes: Nome do mês (opcional)

    Returns:
        Arquivo XLSX para download
    """
    padrao_str = request.args.get('padrao', '')
    mes_nome = request.args.get('mes', '')

    if not padrao_str:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetro "padrao" é obrigatório'
        }), 400

    try:
        # Verificar viabilidade
        if not GeradorPadroesCompletoService.verificar_padrao_viavel(padrao_str):
            return jsonify({
                'sucesso': False,
                'erro': f'Padrão "{padrao_str}" não é viável'
            }), 400

        # Gerar XLSX com viabilidade
        output = GeradorPadroesCompletoService.exportar_para_xlsx_completo(
            padrao_str,
            mes_nome=mes_nome,
            incluir_viabilidade=True
        )

        if output is None:
            return jsonify({
                'sucesso': False,
                'erro': 'Erro ao gerar arquivo XLSX (openpyxl não disponível)'
            }), 500

        padrao_nome = padrao_str.replace(' ', '')

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=padrao_{padrao_nome}_viabilidade.xlsx',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
        )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


# =========================================================================
# API: EXPORTAÇÃO COM SELEÇÃO DE JOGOS
# =========================================================================

@gerador_padroes_bp.route('/api/gerador-padroes/exportar-xlsx-selecionados', methods=['POST'])
def exportar_xlsx_selecionados():
    """
    Exporta jogos selecionados em formato XLSX.

    Body JSON:
        padrao: String do padrão (obrigatório)
        mes: Nome do mês (opcional)
        jogos: Lista de jogos selecionados [{index, numeros}, ...]
        total_gerados: Total de jogos gerados

    Returns:
        Arquivo XLSX para download
    """
    try:
        dados = request.get_json() or {}
        padrao_str = dados.get('padrao', '')
        mes_nome = dados.get('mes', '')
        jogos = dados.get('jogos', [])
        total_gerados = dados.get('total_gerados', 0)

        if not padrao_str:
            return jsonify({
                'sucesso': False,
                'erro': 'Parâmetro "padrao" é obrigatório'
            }), 400

        if not jogos or len(jogos) == 0:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum jogo selecionado para exportar'
            }), 400

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime

            wb = Workbook()
            ws = wb.active
            ws.title = "Apostas Selecionadas"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="D4B31A", end_color="D4B31A", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Informações de cabeçalho
            ws['A1'] = 'Sistema Dia de Sorte - Gerador de Padrões'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:I1')

            ws['A2'] = f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
            ws.merge_cells('A2:D2')

            ws['E2'] = f'Padrão: {padrao_str}'
            ws.merge_cells('E2:I2')

            ws['A3'] = f'Jogos Exportados: {len(jogos)} de {total_gerados}'
            ws['A3'].font = Font(bold=True, color="228B22")
            ws.merge_cells('A3:D3')

            if mes_nome:
                ws['E3'] = f'Mês da Sorte: {mes_nome}'
                ws.merge_cells('E3:I3')

            # Headers da tabela (linha 5)
            headers = ['#', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'Padrão']
            if mes_nome:
                headers.append('Mês')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Dados dos jogos
            for row_idx, jogo in enumerate(jogos, 6):
                numeros = jogo.get('numeros', [])
                index_original = jogo.get('index', row_idx - 5)

                ws.cell(row=row_idx, column=1, value=index_original).border = thin_border
                for i, num in enumerate(numeros[:7], 2):
                    cell = ws.cell(row=row_idx, column=i, value=num)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                ws.cell(row=row_idx, column=9, value=padrao_str).border = thin_border

                if mes_nome:
                    ws.cell(row=row_idx, column=10, value=mes_nome).border = thin_border

            # Ajustar larguras das colunas
            ws.column_dimensions['A'].width = 10
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 6
            ws.column_dimensions['I'].width = 18
            if mes_nome:
                ws.column_dimensions['J'].width = 12

            # Salvar em buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            padrao_nome = padrao_str.replace(' ', '')
            sufixo = f'_{len(jogos)}jogos' if len(jogos) < total_gerados else ''

            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename=padrao_{padrao_nome}{sufixo}.xlsx',
                    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                }
            )

        except ImportError:
            return jsonify({
                'sucesso': False,
                'erro': 'Biblioteca openpyxl não disponível'
            }), 500

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500
