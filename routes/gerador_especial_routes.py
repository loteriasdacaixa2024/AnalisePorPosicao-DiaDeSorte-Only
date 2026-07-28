from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from io import BytesIO
from services.gerador_especial_service import GeradorEspecialService
from services.gerador_digitos_service import GeradorDigitosService
from services.gerador_fatiamento_service import GeradorFatiamentoService
from openpyxl import Workbook
from datetime import datetime
from services.gerador_precisao_service import GeradorPrecisaoService
from services.gerador_atraso_posicao_service import GeradorAtrasoPosicaoService
from services.gerador_atraso_posicao_experimental_service import GeradorAtrasoPosicaoExperimentalService
from services.busca_melhores_saltos_service import BuscaMelhoresSaltosService

gerador_especial_bp = Blueprint('gerador_especial', __name__, url_prefix='/gerador-especial')


@gerador_especial_bp.route('/', methods=['GET'])
def index():
    """Página principal do Gerador Especial."""
    from models.sorteio import Sorteio
    ultimo_sorteio = GeradorEspecialService.get_last_draw()
    stats_mes = GeradorEspecialService.get_month_stats()
    
    concursos_db = Sorteio.query.with_entities(Sorteio.concurso).order_by(Sorteio.concurso.desc()).all()
    concursos_list = [c[0] for c in concursos_db]
    
    return render_template(
        'gerador_especial.html',
        ultimo_sorteio=ultimo_sorteio,
        stats_mes=stats_mes,
        concursos=concursos_list
    )


@gerador_especial_bp.route('/api/gerar', methods=['POST'])
def gerar_apostas():
    """API para gerar apostas."""
    try:
        data = request.json
        
        mes_selecionado = data.get('mes')
        quantidade = int(data.get('quantidade', 10))
        dezenas_por_jogo = int(data.get('dezenas_por_jogo', 7))
        faixa = data.get('faixa', 'livre')
        paridade = data.get('paridade', 'livre')
        temperatura = data.get('temperatura', 'livre')
        valor_aposta = float(data.get('valor_aposta', 2.50))
        
        # Filtros adicionais
        desdobrar_ultimo = data.get('desdobrar_ultimo', False)
        usar_nucleo = data.get('usar_nucleo', False)
        numeros_nucleo = data.get('numeros_nucleo', '')
        ordenacao = data.get('ordenacao', 'crescente')

        # Tratamento do Mês
        if not mes_selecionado or mes_selecionado == '':
            return jsonify({
                'sucesso': False,
                'mensagem': 'Por favor, selecione o mês. Para que as apostas sejam geradas, é preciso selecionar o mês.'
            }), 400
            
        if mes_selecionado == 'aleatorio':
            import random
            mes_selecionado = random.randint(1, 12)
        else:
            try:
                mes_selecionado = int(mes_selecionado)
            except ValueError:
                return jsonify({'sucesso': False, 'mensagem': 'Mês inválido.'}), 400
        
        # Gerar apostas
        resultado = GeradorEspecialService.generate_apostas(
            quantidade, dezenas_por_jogo, mes_selecionado,
            faixa, paridade, temperatura,
            desdobrar_ultimo, usar_nucleo, numeros_nucleo, ordenacao
        )
        
        if not resultado['sucesso']:
            return jsonify(resultado), 400
        
        apostas = resultado['apostas']
        valor_total = quantidade * valor_aposta
        
        return jsonify({
            'sucesso': True,
            'apostas': apostas,
            'quantidade': len(apostas),
            'valor_unitario': valor_aposta,
            'valor_total': valor_total,
            'mes': resultado['mes'],
            'mes_num': resultado['mes_num'],
            'estrategias': {
                'faixa': faixa,
                'paridade': paridade,
                'temperatura': temperatura
            }
        })
    
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/exportar/txt', methods=['POST'])
def exportar_txt():
    """Exporta apostas em TXT."""
    try:
        data = request.json
        apostas = data.get('apostas', [])
        mes_input = data.get('mes', 'JAN')
        
        # Garantir agbreviação se vier número ou nome completo (opcional, mas seguro)
        # O frontend deve enviar já certo, mas vamos garantir.
        # Se for número string '1', converter.
        mes_abreviado = mes_input
        if str(mes_input).isdigit():
            mes_num = int(mes_input)
            mes_abreviado = GeradorEspecialService.MONTHS_ABBR.get(mes_num, 'JAN')
        
        conteudo = GeradorEspecialService.exportar_txt(apostas, mes_abreviado)
        
        buffer = BytesIO()
        buffer.write(conteudo.encode('utf-8'))
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='text/plain',
            as_attachment=True,
            download_name='apostas.txt'
        )
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/exportar/xlsx', methods=['POST'])
def exportar_xlsx():
    """Exporta apostas em XLSX."""
    try:
        data = request.json
        apostas = data.get('apostas', [])
        mes_nome = data.get('mes_nome', 'Janeiro')
        mes_num = int(data.get('mes_num', 1))
        faixa = data.get('faixa', 'livre')
        paridade = data.get('paridade', 'livre')
        temperatura = data.get('temperatura', 'livre')
        valor_unitario = float(data.get('valor_unitario', 2.50))
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Apostas'
        
        # Cabeçalho
        ws['A1'] = 'Concurso'
        ws['B1'] = 'Data'
        ws['C1'] = 'Mês'
        ws['D1'] = 'Faixa'
        ws['E1'] = 'Paridade'
        ws['F1'] = 'Temperatura'
        ws['G1'] = 'Valor Unitário'
        ws['H1'] = 'Valor Total'
        
        ws['A2'] = 'N/A'
        ws['B2'] = datetime.now().strftime('%d/%m/%Y')
        ws['C2'] = mes_nome
        ws['D2'] = faixa
        ws['E2'] = paridade
        ws['F2'] = temperatura
        ws['G2'] = valor_unitario
        ws['H2'] = len(apostas) * valor_unitario
        
        # Cabeçalho de apostas
        for i in range(1, 8):
            ws[f'{chr(71+i)}1'] = f'D{i}'
        
        # Dados de apostas
        for idx, aposta in enumerate(apostas, 3):
            for col_idx, num in enumerate(aposta, 0):
                ws[f'{chr(65+col_idx)}{idx}'] = num
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='apostas.xlsx'
        )
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/exportar/html', methods=['POST'])
def exportar_html():
    """Exporta apostas em HTML."""
    try:
        data = request.json
        apostas = data.get('apostas', [])
        mes_nome = data.get('mes_nome', 'Janeiro')
        faixa = data.get('faixa', 'livre')
        paridade = data.get('paridade', 'livre')
        temperatura = data.get('temperatura', 'livre')
        valor_total = float(data.get('valor_total', 0))
        
        html = GeradorEspecialService.exportar_html(
            apostas, mes_nome, faixa, paridade, temperatura, valor_total
        )
        
        buffer = BytesIO()
        buffer.write(html.encode('utf-8'))
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='text/html',
            as_attachment=True,
            download_name='apostas.html'
        )
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/meses-stats', methods=['GET'])
def meses_stats():
    """Retorna mês mais atrasado e mais frequente."""
    stats = GeradorEspecialService.get_month_stats()
    return jsonify(stats)


@gerador_especial_bp.route('/api/gerar_independente', methods=['POST'])
def gerar_independente():
    """API para o Gerador Independente Avançado (Aba 2)."""
    try:
        data = request.json
        mes_input = data.get('mes', 'aleatorio')
        quantidade = int(data.get('quantidade', 10))
        
        # Filtros Estratégicos
        f_dig = data.get('f_dig', '')
        f_frq = data.get('f_frq', '')
        f_par = data.get('f_par', '')
        f_som = data.get('f_som', '')
        f_seq = data.get('f_seq', '')
        f_rep = data.get('f_rep', '')
        f_lin = data.get('f_lin', '')
        
        # Auxiliares
        quentes = data.get('quentes', [])
        frias = data.get('frias', [])
        ultimo_sorteio = data.get('ultimo_sorteio', [])
        
        usar_origem = bool(data.get('usar_origem', True))
        
        import random
        import math
        
        # Recupera as dezenas autorizadas a jogar (Pool)
        pool_str = data.get('pool_dezenas', [])
        if not pool_str:
            pool = list(range(1, 32))
        else:
            pool = [int(x) for x in pool_str]
            if len(pool) < 7:
                return jsonify({'sucesso': False, 'mensagem': 'A pool de dezenas selecionada precisa ter pelo menos 7 números!'}), 400

        meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        apostas = []
        
        combinacoes_tentadas = 0
        MAX_TENTATIVAS = 600000
        
        def get_combinadic_index(comb):
            idx = 1
            sorted_comb = sorted(comb)
            for i, c in enumerate(sorted_comb):
                start = 1 if i == 0 else sorted_comb[i-1] + 1
                for v in range(start, c):
                    idx += math.comb(31 - v, 6 - i)
            return idx

        while len(apostas) < quantidade and combinacoes_tentadas < MAX_TENTATIVAS:
            combinacoes_tentadas += 1
            
            # GERADOR INDEPENDENTE PURO - Sorteia 7 dezenas a partir do POOL recebido
            # O random.sample garante 7 únicas daquele Pool.
            candidate = random.sample(pool, 7)
            candidate_sorted = sorted(candidate)
            
            # ==== APLICAÇÃO DOS FILTROS ESTRATÉGICOS ====
            
            # 1. Digitos Únicos (ex: '7')
            if f_dig and f_dig != '' and f_dig != 'manual':
                maxDig = int(f_dig)
                digs = set()
                for n in candidate_sorted:
                    s = f"{n:02d}"
                    digs.add(s[0])
                    digs.add(s[1])
                if len(digs) != maxDig:
                    continue
                    
            # 2. Pares e Impares (ex: '4P-3I')
            if f_par and f_par != '' and f_par != 'manual' and 'P' in f_par:
                maxP = int(f_par.split('P')[0])
                maxI = int(f_par.split('-')[1].split('I')[0])
                p = sum(1 for n in candidate_sorted if n % 2 == 0)
                i_cnt = 7 - p
                if p != maxP or i_cnt != maxI:
                    continue
                    
            # 3. Somas (ex: '110 a 119')
            if f_som and f_som != '' and f_som != 'manual' and ' a ' in f_som:
                parts = f_som.split(' a ')
                minS = int(parts[0])
                maxS = int(parts[1])
                s = sum(candidate_sorted)
                # O dropdown original era por dezena ex: 110 a 119. A aposta final tem que cair nessa faixa exata!
                if s < minS or s > maxS:
                    continue
                    
            # 4. Sequencias (ex: '2')
            if f_seq and f_seq != '' and f_seq != 'manual':
                maxS = int(f_seq)
                mSeq = 1
                curr = 1
                for j in range(1, 7):
                    if candidate_sorted[j] == candidate_sorted[j-1] + 1:
                        curr += 1
                    else:
                        mSeq = max(mSeq, curr)
                        curr = 1
                mSeq = max(mSeq, curr)
                if mSeq != maxS:
                    continue
                    
            # 5. Distância de Linhas (ex: '3-2-1-1')
            if f_lin and f_lin != '' and f_lin != 'manual' and '-' in f_lin:
                l = [0, 0, 0, 0]
                for n in candidate_sorted:
                    if n <= 10: l[0] += 1
                    elif n <= 20: l[1] += 1
                    elif n <= 30: l[2] += 1
                    else: l[3] += 1
                # Formatando a linha sorteada em ex: '3-2-1-1'
                l_filtered = sorted([x for x in l if x > 0], reverse=True)
                lin_str = '-'.join(map(str, l_filtered))
                if lin_str != f_lin:
                    continue
                    
            # 6. Repetições do concurso anterior (ex: '2')
            if f_rep and f_rep != '' and f_rep != 'manual':
                maxRep = int(f_rep)
                rep = sum(1 for n in candidate_sorted if n in ultimo_sorteio)
                if rep != maxRep:
                    continue
                    
            # 7. Frequencia (ex: '3Q-2M-2F')
            if f_frq and f_frq != '' and f_frq != 'manual' and 'Q' in f_frq:
                qMax = int(f_frq.split('Q')[0])
                mMax = int(f_frq.split('-')[1].split('M')[0])
                fMax = int(f_frq.split('-')[2].split('F')[0])
                q = m = f = 0
                for n in candidate_sorted:
                    if n in quentes: q += 1
                    elif n in frias: f += 1
                    else: m += 1
                if q != qMax or m != mMax or f != fMax:
                    continue
                    
            # PASXOU EM TODOS OS FILTROS ESTRATÉGICOS!
            
            mes_num = random.randint(1, 12) if str(mes_input) == 'aleatorio' else int(mes_input)
            linha_id = get_combinadic_index(candidate_sorted) if usar_origem else None

            apostas.append({
                'dezenas_p': candidate,          # Unordered draws array
                'dezenas': candidate_sorted,     # Ordered to comply with frontend/export APIs
                'mes_num': mes_num,
                'mes_nome': meses_nomes[mes_num],
                'linha_original': linha_id
            })
            
        if len(apostas) == 0:
            return jsonify({'sucesso': False, 'mensagem': 'Nenhuma aposta encontrada! Combinação dos filtros impossível ou muito restritiva.'})

            
        return jsonify({
            'sucesso': True,
            'quantidade': len(apostas),
            'apostas': apostas
        })
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)})


@gerador_especial_bp.route('/api/digitos-frequencia', methods=['GET'])
def digitos_frequencia():
    """Retorna o histórico de frequências reais dos dígitos na loteria."""
    try:
        data = GeradorDigitosService.get_digit_frequencies()
        return jsonify({'sucesso': True, 'frequencias': data})
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/digitos-top-combinacoes', methods=['GET'])
def digitos_top_combinacoes():
    """Retorna as top combinacoes de dígitos."""
    try:
        data = GeradorDigitosService.get_top_combinations(3)
        return jsonify({'sucesso': True, 'top_combinacoes': data})
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/gerar_por_digitos', methods=['POST'])
def gerar_por_digitos():
    try:
        data = request.json
        digitos = data.get('digitos', [])
        qtd_aposta = int(data.get('quantidade_apostas', 10))
        dezenas_por_jogo = int(data.get('dezenas_por_jogo', 7))
        excluidas = data.get('dezenas_excluidas', [])
        mes_input = data.get('mes_input', 'aleatorio')
        
        # Get Month Stats if needed
        if mes_input in ['frequente', 'atrasado']:
            stats = GeradorEspecialService.get_month_stats()
            if mes_input == 'frequente':
                mes_input = stats['frequente']
            else:
                mes_input = stats['atrasado']
        
        filtros = {
            'f_frq': data.get('f_frq', ''),
            'f_par': data.get('f_par', ''),
            'f_som': data.get('f_som', ''),
            'f_seq': data.get('f_seq', ''),
            'f_rep': data.get('f_rep', ''),
            'f_lin': data.get('f_lin', ''),
            'quentes': data.get('quentes', []),
            'frias': data.get('frias', []),
            'ultimo_sorteio': data.get('ultimo_sorteio', [])
        }
        
        # Validations Min 5 digits, Max 9 digits
        if len(digitos) < 5 or len(digitos) > 9:
            return jsonify({'sucesso': False, 'mensagem': 'Você deve selecionar entre 5 e 9 dígitos!'}), 400
            
        resultado = GeradorDigitosService.gerar_apostas_por_digitos(
            digitos, dezenas_por_jogo, qtd_aposta, excluidas, mes_input, filtros=filtros
        )
        
        if not resultado['sucesso']:
            return jsonify(resultado), 400
            
        # Return the exact list of dicts that GeradorDigitosService provided
        apostas = resultado['apostas']
            
        return jsonify({
            'sucesso': True,
            'quantidade': len(apostas),
            'apostas': apostas,
            'pool_dezenas': resultado['pool_dezenas'],
            'pool_tamanho': resultado['pool_tamanho'],
            'combinacoes_possiveis': resultado['combinacoes_possiveis']
        })
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/precisao_matriz', methods=['GET'])
def precisao_matriz():
    try:
        resultado = GeradorPrecisaoService.gerar_matriz_precisao()
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/precisao_lotes', methods=['POST'])
def precisao_lotes():
    try:
        data = request.json
        matriz = data.get('matriz', [])
        fechamento_total = data.get('fechamento_total', False)
        
        if not matriz or len(matriz) != 9:
            return jsonify({'sucesso': False, 'mensagem': 'É necessário fornecer exatamente 9 dezenas na matriz.'}), 400
            
        resultado = GeradorPrecisaoService.gerar_lotes_multiplexados(matriz, fechamento_total)
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/ultimos-resultados', methods=['GET'])
def ultimos_resultados():
    """Retorna os últimos concursos para o Laboratório de Backtesting."""
    try:
        from models.sorteio import Sorteio
        sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).all()
        dados = []
        for s in sorteios:
            dados.append({
                'concurso': s.concurso,
                'dezenas': s.get_posicoes_lista(),
                'mes_nome': s.get_nome_mes(),
                'mes_num': s.mes_sorte
            })
        return jsonify({'sucesso': True, 'sorteios': dados})
    except Exception as e:
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/estrategico_gerar', methods=['POST'])
def estrategico_gerar():
    try:
        from models.sorteio import Sorteio
        # Get all draws ordered
        sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).limit(30).all()
        if not sorteios:
            return jsonify({'sucesso': False, 'mensagem': 'Sem sorteios no banco!'})
            
        last_concurso = sorteios[0].concurso
        
        # Calculate hot/medium/delayed using similar logic as script
        freq = {i: 0 for i in range(1, 32)}
        last_seen = {i: 0 for i in range(1, 32)}
        
        # we really need all draws to calculate accurate delay, or at least last 150
        todos_sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).limit(150).all()
        for s in todos_sorteios:
            c = s.concurso
            nums = s.get_posicoes_lista()
            for n in nums:
                if 1 <= n <= 31:
                    if last_seen[n] == 0:
                        last_seen[n] = c
                        
        delays = {}
        for i in range(1, 32):
            if last_seen[i] > 0:
                delays[i] = last_concurso - last_seen[i]
            else:
                delays[i] = 150
                
        for s in sorteios: # only last 30 for freq
            nums = s.get_posicoes_lista()
            for n in nums:
                if 1 <= n <= 31:
                    freq[n] += 1
                    
        # Define Atrasadas: top 5 by delay
        sorted_delay = sorted(delays.items(), key=lambda x: x[1], reverse=True)
        atrasadas = [x[0] for x in sorted_delay[:5]]
        
        # Quentes: top 7 by freq, that are NOT in atrasadas
        sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)
        quentes = []
        for x in sorted_freq:
            if x[0] not in atrasadas:
                quentes.append(x[0])
            if len(quentes) >= 7:
                break
                
        # Medias: all other
        medias = [i for i in range(1, 32) if i not in atrasadas and i not in quentes]
        
        import random
        apostas = []
        
        patterns = [
            (3, 3, 1), (4, 2, 1), (3, 3, 1), (4, 2, 1), (3, 3, 1),
            (3, 3, 1), (4, 2, 1), (3, 3, 1), (4, 2, 1), (3, 3, 1)
        ]
        
        meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        # balance atrasadas (5 atrasadas * 2 = 10 picks)
        atrasadas_queue = atrasadas * 2
        random.shuffle(atrasadas_queue)
        
        used_patterns = set()
        
        for idx, (h, m, d) in enumerate(patterns):
            my_d = [atrasadas_queue[idx]]
            best_bet = None
            
            # Tentar até 150 vezes para encontrar um padrão estrutural único (ex: 0011223)
            for attempt in range(150):
                my_h = random.sample(quentes, h)
                my_m = random.sample(medias, m)
                candidate = sorted(my_h + my_m + my_d)
                
                # Identifica o Padrão Inicial (0 a 3)
                pattern_str = "".join([str(n // 10) for n in candidate])
                
                if pattern_str not in used_patterns:
                    best_bet = candidate
                    used_patterns.add(pattern_str)
                    break
            
            if not best_bet:
                # Se não encontrar num limite de 150 tentativas, pega qualquer um
                my_h = random.sample(quentes, h)
                my_m = random.sample(medias, m)
                best_bet = sorted(my_h + my_m + my_d)
            
            mes_num = random.randint(1, 12)
            while mes_num in best_bet:
                mes_num = random.randint(1, 12)
                
            apostas.append({
                'dezenas': best_bet,
                'mes_num': mes_num,
                'mes_nome': meses_nomes[mes_num]
            })
            
        # Ordenar as apostas matematicamente na vertical para melhor leitura estrutural
        apostas.sort(key=lambda x: x['dezenas'])
            
        return jsonify({
            'sucesso': True,
            'apostas': apostas,
            'atrasadas': atrasadas,
            'quentes': quentes,
            'medias': medias
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500

@gerador_especial_bp.route('/api/enviar-conferencia', methods=['POST'])
def enviar_especial_conferencia():
    import os
    from models.conferencia_historica import db
    from services.conferencia_historica_service import ConferenciaHistoricaService
    
    data = request.json
    concurso_base = data.get('concurso_base')
    matriz_original = data.get('matriz_original')
    matriz_ajustada = data.get('matriz_ajustada')
    
    if not matriz_original:
        return jsonify({'sucesso': False, 'erro': 'Matriz original não fornecida'}), 400
        
    afin_dir = os.path.join(os.getcwd(), 'conferencia_filtros-baixados')
    os.makedirs(afin_dir, exist_ok=True)
    
    is_sinc = (concurso_base == 'Sincronizado_Exclusao')
    detalhes = data.get('detalhes', '')
    
    if is_sinc:
        import datetime
        agora = datetime.datetime.now().strftime("%d%m%H%M")
        nome_original = f"Aba7_Sincronizado_{agora}.txt"
        desc_original = f"Aba 7 (Sincronizado) - {detalhes}" if detalhes else "Fechamento Sincronizado"
    elif concurso_base == 'Estatisticas_Ciclo_Por_Posicao':
        import datetime
        agora = datetime.datetime.now().strftime("%d%m%H%M%S")
        nome_original = f"Ciclo_Por_Posicao_{agora}.txt"
        desc_original = detalhes or 'Matriz Ciclos Faltantes — Estatísticas'
    else:
        nome_original = f"Gerador_Atraso_Posicional_{concurso_base}.txt"
        desc_original = f"Duelo: Especial Posicional ({nome_original})"
    
    path_original = os.path.join(afin_dir, nome_original)
    with open(path_original, 'w') as f: f.write(matriz_original)
        
    mensagens = []
    try:
        sessao_o = ConferenciaHistoricaService.criar_sessao(
            nome_arquivo=nome_original,
            descricao=desc_original,
            estrategia='ordenada', filtro_min=4
        )
        resultado_proc = ConferenciaHistoricaService.processar_arquivo(
            sessao_o.id, matriz_original
        )
        if not resultado_proc.get('sucesso'):
            return jsonify(
                {
                    'sucesso': False,
                    'erro': resultado_proc.get('erro', 'Falha ao processar matriz'),
                    'sessao_id': sessao_o.id,
                }
            ), 400
        db.session.refresh(sessao_o)
        mensagens.append(f"Matriz enviada (ID: {sessao_o.id})")
        
        if matriz_ajustada:
            nome_ajustada = f"Gerador_Atraso_Ajustado_{concurso_base}.txt"
            path_ajustada = os.path.join(afin_dir, nome_ajustada)
            with open(path_ajustada, 'w') as f: f.write(matriz_ajustada)
            
            sessao_a = ConferenciaHistoricaService.criar_sessao(
                nome_arquivo=nome_ajustada,
                descricao=f"Duelo: Especial Ajustado ({nome_ajustada})",
                estrategia='ordenada', filtro_min=4
            )
            ConferenciaHistoricaService.processar_arquivo(sessao_a.id, matriz_ajustada)
            mensagens.append(f"Ajustado enviado (ID: {sessao_a.id})")
        
        return jsonify({
            'sucesso': True,
            'mensagem': "Arquivos enviados para a Central de Conferência:\n" + "\n".join(mensagens),
            'sessao_id': sessao_o.id,
            'status': sessao_o.status,
            'total_apostas': sessao_o.total_apostas,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@gerador_especial_bp.route('/api/enviar-batalha-conferencia', methods=['POST'])
def enviar_batalha_conferencia():
    import os
    import datetime
    from services.conferencia_historica_service import ConferenciaHistoricaService
    
    data = request.json
    resultados = data.get('batalha_resultados', [])
    
    if not resultados:
        return jsonify({'sucesso': False, 'erro': 'Nenhum resultado enviado.'}), 400
        
    afin_dir = os.path.join(os.getcwd(), 'conferencia_filtros-baixados')
    os.makedirs(afin_dir, exist_ok=True)
    
    agora = datetime.datetime.now().strftime("%d%m%H%M")
    
    mensagens = []
    try:
        for res in resultados:
            nome_gerador = res.get('nome', 'Desconhecido')
            apostas = res.get('apostas', [])
            
            if not apostas:
                continue
                
            # Converter array de apostas para formato TXT (matriz_txt)
            matriz_txt = ""
            for ap in apostas:
                if isinstance(ap, dict):
                    dezenas = ap.get('dezenas', [])
                    mes = ap.get('mes_nome', 'Jan')
                elif isinstance(ap, list):
                    dezenas = ap
                    mes = 'Jan'
                else:
                    continue
                    
                try:
                    nums = " ".join([f"{int(n):02d}" for n in dezenas])
                except ValueError:
                    # Fallback in case n cannot be cast to int
                    nums = " ".join([str(n) for n in dezenas])
                    
                matriz_txt += f"{nums} {mes}\r\n"
                
            nome_arquivo = f"Batalha_{nome_gerador.replace(' ', '_')}_{agora}.txt"
            desc_original = f"[Batalha] {nome_gerador}"
            
            path_original = os.path.join(afin_dir, nome_arquivo)
            with open(path_original, 'w') as f:
                f.write(matriz_txt)
                
            sessao = ConferenciaHistoricaService.criar_sessao(
                nome_arquivo=nome_arquivo,
                descricao=desc_original,
                estrategia='ordenada', filtro_min=4
            )
            ConferenciaHistoricaService.processar_arquivo(sessao.id, matriz_txt)
            mensagens.append(f"Criada sessão para {nome_gerador} (ID: {sessao.id})")
            
        return jsonify({'sucesso': True, 'mensagem': "Sessões criadas:\n" + "\n".join(mensagens)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@gerador_especial_bp.route('/api/gerar_atraso_posicao', methods=['POST'])
def gerar_atraso_posicao():
    try:
        data = request.json
        quantidade = int(data.get('quantidade', 10))
        dezenas_por_jogo = int(data.get('dezenas_por_jogo', 7))
        mes_selecionado = data.get('mes_tipo', 'aleatorio')
        concurso_base_id = data.get('concurso_base_id', 'ultimo')

        resultado = GeradorAtrasoPosicaoService.gerar_apostas_atraso_posicao(
            concurso_base_id=concurso_base_id,
            quantidade=quantidade,
            dezenas_por_jogo=dezenas_por_jogo,
            mes_selecionado=mes_selecionado
        )

        return jsonify(resultado)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/gerar_atraso_posicao_experimental', methods=['POST'])
def gerar_atraso_posicao_experimental():
    try:
        data = request.json or {}
        resultado = GeradorAtrasoPosicaoExperimentalService.gerar_apostas_atraso_posicao_experimental(
            concurso_base_id=data.get('concurso_base_id', 'ultimo'),
            quantidade=int(data.get('quantidade', 0)),
            dezenas_por_jogo=int(data.get('dezenas_por_jogo', 7)),
            mes_selecionado=data.get('mes_tipo', 'sequencial'),
            salto_modo=data.get('salto_modo', 'global'),
            salto_global=int(data.get('salto_global', 1)),
            salto_global_menos=data.get('salto_global_menos'),
            salto_simetrico=data.get('salto_simetrico', True),
            saltos_coluna=data.get('saltos_coluna'),
            saltos_coluna_menos=data.get('saltos_coluna_menos'),
            limite_salto_max=int(data.get('limite_salto_max', 30)),
        )
        return jsonify(resultado)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


# ---------------------------------------------------------------------------
# Busca automática dos melhores saltos (Aba 6A) — complementar, não altera o gerador
# ---------------------------------------------------------------------------

@gerador_especial_bp.route('/api/busca_melhores_saltos/iniciar', methods=['POST'])
def busca_melhores_saltos_iniciar():
    try:
        data = request.json or {}
        params = {
            'concurso_base_id': data.get('concurso_base_id', 'ultimo'),
            'dezenas_por_jogo': int(data.get('dezenas_por_jogo', 7)),
            'mes_tipo': data.get('mes_tipo', 'sequencial'),
            'limite_salto_max': int(data.get('limite_salto_max', 6)),
            'max_testes': int(data.get('max_testes', 1000)),
            'simetrico': bool(data.get('simetrico', True)),
            'modo_busca': data.get('modo_busca', 'aleatorio'),
            'seed': data.get('seed'),
            'usar_ajustada': bool(data.get('usar_ajustada', False)),
            'janela_concursos': int(data.get('janela_concursos', 200)),
            'top_n': int(data.get('top_n', 50)),
            'escopo_percurso': data.get('escopo_percurso', 'todos'),
        }
        if params['limite_salto_max'] not in (6, 15, 30):
            params['limite_salto_max'] = 6
        params['max_testes'] = max(1, min(params['max_testes'], 20000))
        params['top_n'] = max(10, min(params['top_n'], 50))

        app = current_app._get_current_object()
        job_id = BuscaMelhoresSaltosService.iniciar_busca_background(app, params)
        return jsonify({'sucesso': True, 'job_id': job_id})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/busca_melhores_saltos/status/<job_id>', methods=['GET'])
def busca_melhores_saltos_status(job_id):
    job = BuscaMelhoresSaltosService.obter_job(job_id)
    if not job:
        return jsonify({'sucesso': False, 'mensagem': 'Job não encontrado.'}), 404
    return jsonify({'sucesso': True, **job})


@gerador_especial_bp.route('/api/busca_melhores_saltos/cancelar/<job_id>', methods=['POST'])
def busca_melhores_saltos_cancelar(job_id):
    ok = BuscaMelhoresSaltosService.cancelar_job(job_id)
    if not ok:
        return jsonify({'sucesso': False, 'mensagem': 'Job não encontrado.'}), 404
    return jsonify({'sucesso': True})


@gerador_especial_bp.route('/api/busca_melhores_saltos/apostas/<job_id>/<int:rank>', methods=['GET'])
def busca_melhores_saltos_apostas(job_id, rank):
    blob = BuscaMelhoresSaltosService.obter_apostas_rank(job_id, rank)
    if not blob:
        return jsonify({'sucesso': False, 'mensagem': 'Apostas não encontradas para este rank.'}), 404
    return jsonify({'sucesso': True, 'rank': rank, **blob})


@gerador_especial_bp.route('/api/busca_melhores_saltos/exportar', methods=['POST'])
def busca_melhores_saltos_exportar():
    """Exporta exatamente as apostas já encontradas (não regenera)."""
    try:
        data = request.json or {}
        job_id = data.get('job_id')
        rank = int(data.get('rank', 1))
        formato = (data.get('formato') or 'txt').lower()
        blob = BuscaMelhoresSaltosService.obter_apostas_rank(job_id, rank)
        if not blob:
            return jsonify({'sucesso': False, 'mensagem': 'Apostas não encontradas.'}), 404

        apostas = blob.get('apostas') or []
        meses_abbr = {
            1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez',
        }
        concurso = blob.get('concurso_base', 'X')
        preset_safe = (blob.get('preset') or 'preset').replace('/', '-').replace(' ', '')[:40]

        if formato == 'csv':
            linhas = ['dezenas;mes']
            for ap in apostas:
                nums = ' '.join(f'{n:02d}' for n in (ap.get('numeros') or []))
                mes = meses_abbr.get(ap.get('mes_num'), ap.get('mes_nome') or '')
                linhas.append(f'{nums};{mes}')
            conteudo = '\r\n'.join(linhas) + '\r\n'
            buffer = BytesIO(conteudo.encode('utf-8-sig'))
            return send_file(
                buffer,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'BuscaSaltos_Rank{rank}_C{concurso}_{preset_safe}.csv',
            )

        if formato == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Apostas'
            ws.append(['Linha', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'Mes', 'Preset', 'Rank'])
            for ap in apostas:
                nums = list(ap.get('numeros') or [])
                while len(nums) < 7:
                    nums.append('')
                mes = meses_abbr.get(ap.get('mes_num'), ap.get('mes_nome') or '')
                ws.append(
                    [ap.get('linha_offset')]
                    + nums[:15]
                    + [mes, blob.get('preset'), rank]
                )
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'BuscaSaltos_Rank{rank}_C{concurso}_{preset_safe}.xlsx',
            )

        # TXT (padrão 6A)
        linhas = []
        for ap in apostas:
            nums = ' '.join(f'{n:02d}' for n in (ap.get('numeros') or []))
            mes = meses_abbr.get(ap.get('mes_num'), ap.get('mes_nome') or '')
            linhas.append(f'{nums} {mes}')
        conteudo = '\r\n'.join(linhas) + '\r\n'
        buffer = BytesIO(conteudo.encode('utf-8'))
        return send_file(
            buffer,
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'BuscaSaltos_Rank{rank}_C{concurso}_{preset_safe}.txt',
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/fechamento-sincronizado', methods=['POST'])
def fechamento_sincronizado():
    try:
        from models.sorteio import Sorteio
        import itertools
        from services.analise_meses_service import AnaliseMesesService
        
        data = request.json
        k_grupos = int(data.get('grupos', 6))
        valor_unitario = float(data.get('valor_unitario', 2.50))
        
        # 1. Pegar o último concurso
        ultimo_sorteio_db = Sorteio.query.order_by(Sorteio.concurso.desc()).first()
        if not ultimo_sorteio_db:
            return jsonify({'sucesso': False, 'mensagem': 'Nenhum sorteio encontrado no banco de dados.'})
            
        dezenas_ultimo = ultimo_sorteio_db.get_posicoes_lista()
        mes_atual = ultimo_sorteio_db.mes_sorte if ultimo_sorteio_db else 1
        
        # Obter meses estratégicos dinamicamente
        estats_meses = AnaliseMesesService.obter_estatisticas_meses()
        meses_list = estats_meses['meses']
        
        meses_freq = sorted(meses_list, key=lambda x: x['frequencia'], reverse=True)
        
        meses_estrategicos = [mes_atual]
        
        # Garantir 3 mais atrasados (totalizando 4 meses com o atual)
        for m in meses_list:
            if m['numero'] not in meses_estrategicos:
                meses_estrategicos.append(m['numero'])
            if len(meses_estrategicos) == 4:
                break
                
        # Garantir 3 mais frequentes (totalizando 7 meses distintos)
        for m in meses_freq:
            if m['numero'] not in meses_estrategicos:
                meses_estrategicos.append(m['numero'])
            if len(meses_estrategicos) == 7:
                break
        
        # 2. Excluir as 7 dezenas sorteadas
        dezenas_restantes = [n for n in range(1, 32) if n not in dezenas_ultimo]
        
        if len(dezenas_restantes) != 24:
            return jsonify({'sucesso': False, 'mensagem': f'Erro: O número de dezenas restantes não é 24 (é {len(dezenas_restantes)}). Base de dados pode ter sorteios inválidos.'})
            
        # 3. Dividir as 24 restantes sequencialmente em 8 grupos de 3 dezenas (G1 a G8)
        dezenas_restantes.sort()
        grupos = []
        for i in range(8):
            g = dezenas_restantes[i*3 : (i+1)*3]
            grupos.append(g)
            
        grupos_exibicao = [{'nome': f'Grupo {i+1}', 'dezenas': grupos[i]} for i in range(8)]
            
        # 4. Reduzir cada grupo de 3 dezenas de 2 em 2 (3 pares) e sincronizar as reduções
        linhas_sincronizadas = [[], [], []]
        for g in grupos:
            linhas_sincronizadas[0].append((g[0], g[1]))
            linhas_sincronizadas[1].append((g[0], g[2]))
            linhas_sincronizadas[2].append((g[1], g[2]))
            
        # 5. Realizar as combinações do número de grupos escolhidos
        indices_grupos = list(range(8))
        combinacoes_grupos = list(itertools.combinations(indices_grupos, k_grupos))
        
        apostas = []
        meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        # 6. Multiplicar cada combinação pelas 3 linhas sincronizadas e alternar meses
        idx_aposta = 0
        for comb in combinacoes_grupos:
            for linha_idx in range(3):
                dezenas_aposta = []
                for grupo_idx in comb:
                    dupla = linhas_sincronizadas[linha_idx][grupo_idx]
                    dezenas_aposta.extend(dupla)
                
                dezenas_aposta.sort()
                
                # Alternar entre os 7 meses dinâmicos
                mes_num = meses_estrategicos[idx_aposta % 7]
                
                apostas.append({
                    'dezenas': dezenas_aposta,
                    'mes_num': mes_num,
                    'mes_nome': meses_nomes[mes_num]
                })
                idx_aposta += 1
                
        # 7. Calcular as métricas financeiras usando o custo da aposta fornecido
        quantidade = len(apostas)
        valor_total = quantidade * valor_unitario
        
        return jsonify({
            'sucesso': True,
            'apostas': apostas,
            'quantidade': quantidade,
            'valor_total': valor_total,
            'grupos_exibicao': grupos_exibicao
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500

@gerador_especial_bp.route('/api/gerar-ciclo-posicional', methods=['POST'])
def gerar_ciclo_posicional():
    try:
        from models.sorteio import Sorteio
        from services.analise_ciclos_dezenas_service import AnaliseCiclosDezenasService
        from services.analise_meses_service import AnaliseMesesService
        from datetime import datetime
        import random
        
        data = request.json
        historico_tamanho = int(data.get('historico_tamanho', 10))
        quantidade = int(data.get('quantidade', 10))
        mes_tipo = data.get('mes_tipo', 'inteligente')
        
        # 1. Obter pendências do ciclo
        ciclo_atual = AnaliseCiclosDezenasService.obter_ciclo_atual()
        dezenas_pendentes = ciclo_atual['dezenas_pendentes'] if ciclo_atual else []
        
        # 2. Obter sorteios para o histórico posicional + o último sorteio (para exclusão)
        sorteios = Sorteio.query.order_by(Sorteio.concurso.desc()).limit(historico_tamanho + 1).all()
        if not sorteios:
            return jsonify({'sucesso': False, 'mensagem': 'Nenhum sorteio encontrado no banco.'})
            
        sorteios.reverse() # Ordena do mais antigo para o mais novo
        
        ultimo_sorteio = sorteios[-1]
        dezenas_ultimo = [
            ultimo_sorteio.posicao_1, ultimo_sorteio.posicao_2, ultimo_sorteio.posicao_3,
            ultimo_sorteio.posicao_4, ultimo_sorteio.posicao_5, ultimo_sorteio.posicao_6, ultimo_sorteio.posicao_7
        ]
        
        # O histórico posicional não inclui o último sorteio
        historico = sorteios[:-1] 
        
        # 3. Excluir dezenas do último sorteio da pool de apostas
        dezenas_validas_ciclo = [d for d in dezenas_pendentes if d not in dezenas_ultimo]
        todas_validas = [d for d in range(1, 32) if d not in dezenas_ultimo]
        
        alerta = None
        if len(dezenas_validas_ciclo) < 7:
            alerta = f"O ciclo possui apenas {len(dezenas_validas_ciclo)} dezenas válidas após a exclusão do último sorteio. As posições vazias serão preenchidas pelas dezenas gerais mais quentes do histórico selecionado."
            
        # 4. Calcular Frequência Posicional (Colunas 1 a 7)
        freq_pos = {p: {d: 0 for d in range(1, 32)} for p in range(1, 8)}
        for s in historico:
            freq_pos[1][s.posicao_1] += 1
            freq_pos[2][s.posicao_2] += 1
            freq_pos[3][s.posicao_3] += 1
            freq_pos[4][s.posicao_4] += 1
            freq_pos[5][s.posicao_5] += 1
            freq_pos[6][s.posicao_6] += 1
            freq_pos[7][s.posicao_7] += 1
            
        # Preparar meses dinâmicos
        stats_meses_data = AnaliseMesesService.obter_estatisticas_meses()
        meses_lista_atraso = [m['numero'] for m in stats_meses_data['meses']]
        meses_lista_freq = [m['numero'] for m in sorted(stats_meses_data['meses'], key=lambda x: x['frequencia'], reverse=True)]
        
        meses_estrategicos = [
            meses_lista_atraso[0] if meses_lista_atraso else 1,
            meses_lista_freq[0] if meses_lista_freq else 1,
            datetime.now().month,
            meses_lista_atraso[1] if len(meses_lista_atraso) > 1 else (meses_lista_atraso[0] if meses_lista_atraso else 1),
            meses_lista_freq[1] if len(meses_lista_freq) > 1 else (meses_lista_freq[0] if meses_lista_freq else 1),
            meses_lista_atraso[2] if len(meses_lista_atraso) > 2 else (meses_lista_atraso[0] if meses_lista_atraso else 1),
            meses_lista_freq[2] if len(meses_lista_freq) > 2 else (meses_lista_freq[0] if meses_lista_freq else 1)
        ]
        meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        # 5. Gerar as Apostas
        apostas = []
        for i in range(quantidade):
            aposta = []
            disponiveis = set(todas_validas)
            pool_ciclo = list(dezenas_validas_ciclo)
            
            # Para criar variação nas apostas, fazemos um embaralhamento ponderado
            # Priorizamos preencher com as dezenas do ciclo primeiro.
            # Vamos iterar pelas posições 1 a 7 e achar o melhor candidato
            for p in range(1, 8):
                # Se ainda há dezenas no pool do ciclo, buscamos os melhores para aquela posição
                candidatos = pool_ciclo if pool_ciclo else list(disponiveis)
                
                # Ordena os candidatos pela frequência na posição 'p'
                candidatos.sort(key=lambda x: freq_pos[p][x], reverse=True)
                
                # Pegar os top 3 para dar variação estatística, ou todos se forem poucos
                melhores = [c for c in candidatos if freq_pos[p][c] > 0][:3]
                
                # Se não houver nenhum "quente" nessa posição, pega aleatório dos candidatos
                if melhores:
                    escolhida = random.choice(melhores)
                else:
                    escolhida = random.choice(candidatos)
                    
                aposta.append(escolhida)
                
                # Remover das listas para não duplicar
                if escolhida in pool_ciclo: 
                    pool_ciclo.remove(escolhida)
                if escolhida in disponiveis:
                    disponiveis.remove(escolhida)
                    
            aposta.sort()
            
            # 6. Seleção do Mês
            mes_num = 1
            if mes_tipo == 'inteligente':
                mes_num = meses_estrategicos[i % 7]
            elif mes_tipo == 'aleatorio':
                mes_num = random.randint(1, 12)
            else:
                try:
                    mes_num = int(mes_tipo)
                except ValueError:
                    mes_num = random.randint(1, 12)
                    
            apostas.append({
                'dezenas': aposta,
                'mes_num': mes_num,
                'mes_nome': meses_nomes[mes_num]
            })
            
        # 7. Montar o relatório visual da análise para mostrar na tela
        analise_tela = []
        for d in dezenas_validas_ciclo:
            frequencias_d = [(p, freq_pos[p][d]) for p in range(1, 8) if freq_pos[p][d] > 0]
            frequencias_d.sort(key=lambda x: x[1], reverse=True)
            
            if frequencias_d:
                top = frequencias_d[:2]
                texto_posicoes = " e ".join([f"Coluna {p} ({f}x)" for p, f in top])
                analise_tela.append(f"O <strong>{d:02d}</strong> tem alta afinidade com a <strong>{texto_posicoes}</strong>")
            else:
                analise_tela.append(f"O <strong>{d:02d}</strong> não apareceu no histórico selecionado.")
            
        return jsonify({
            'sucesso': True,
            'apostas': apostas,
            'dezenas_pendentes': dezenas_validas_ciclo,
            'analise_tela': analise_tela,
            'alerta': alerta
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500


@gerador_especial_bp.route('/api/gerador-especial/gerar-fatiamento', methods=['POST'])
def gerar_fatiamento():
    dados = request.get_json()
    qtd_apostas = int(dados.get('qtd_apostas', 10))
    dezenas_por_jogo = int(dados.get('dezenas_por_jogo', 7))
    filtros = dados.get('filtros', {})
    limites_fatiamento = dados.get('limites', {})
    
    # Tratando limites para ter os defaults em int
    limites = {str(d): int(limites_fatiamento.get(str(d), 7)) for d in range(10)}
    
    resultado = GeradorFatiamentoService.gerar_apostas(
        qtd_apostas=qtd_apostas, 
        dezenas_por_jogo=dezenas_por_jogo, 
        filtros=filtros, 
        limites_fatiamento=limites
    )
    
    return jsonify(resultado)


@gerador_especial_bp.route('/api/hibrido-colunas/analise', methods=['GET'])
def hibrido_colunas_analise():
    """Retorna análise pré-estratégia para a modalidade Dia de Sorte."""
    from services.gerador_hibrido_colunas_service import GeradorHibridoColunasService
    resultado = GeradorHibridoColunasService.get_hibrido_analise()
    return jsonify(resultado)


@gerador_especial_bp.route('/api/hibrido-colunas/gerar', methods=['POST'])
def hibrido_colunas_gerar():
    """Gera apostas baseadas na estratégia Híbrida Estrutural por Colunas."""
    from services.gerador_hibrido_colunas_service import GeradorHibridoColunasService
    try:
        data = request.json or {}
        
        quantidade = int(data.get('quantidade', 10))
        dezenas_por_jogo = int(data.get('dezenas_por_jogo', 7))
        modo = data.get('modo', 'hibrido')
        colunas_selecionadas = data.get('colunas_selecionadas', [])
        fechar_colunas_completas = data.get('fechar_colunas_completas', [])
        reaproveitar_restantes = bool(data.get('reaproveitar_restantes', True))
        cobertura_desejada = data.get('cobertura_desejada', 'alta')
        mes_selecionado = data.get('mes', 'aleatorio')
        valor_aposta = float(data.get('valor_aposta', 2.50))
        
        resultado = GeradorHibridoColunasService.gerar_apostas_hibrido(
            quantidade=quantidade,
            dezenas_por_jogo=dezenas_por_jogo,
            modo=modo,
            colunas_selecionadas=colunas_selecionadas,
            fechar_colunas_completas=fechar_colunas_completas,
            reaproveitar_restantes=reaproveitar_restantes,
            cobertura_desejada=cobertura_desejada,
            mes_selecionado=mes_selecionado,
            valor_aposta=valor_aposta
        )
        
        if not resultado['sucesso']:
            return jsonify(resultado), 400
            
        return jsonify(resultado)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'sucesso': False, 'mensagem': str(e)}), 500

