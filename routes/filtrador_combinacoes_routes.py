# -*- coding: utf-8 -*-
"""
Rotas para o sistema de filtragem de combinações do Dia de Sorte
"""

from flask import Blueprint, render_template, jsonify, request, send_file
from datetime import datetime
import io
import os
from services.filtrador_combinacoes_service import FiltradorCombinacoesService, gerar_arquivo_combinacoes_txt

# Cria o blueprint
filtrador_bp = Blueprint('filtrador_combinacoes', __name__, url_prefix='/filtrador-combinacoes')

# Instância do serviço será inicializada apenas sob demanda dentro do contexto da aplicação!
_servico_filtro = None

def obter_servico():
    global _servico_filtro
    if _servico_filtro is None:
        _servico_filtro = FiltradorCombinacoesService()
    return _servico_filtro

# Estado global para processamento (pausar/continuar)
estado_processamento = {
    'em_andamento': False,
    'pausado': False,
    'progresso': 0,
    'resultado_atual': None
}


@filtrador_bp.route('/')
def index():
    """Renderiza a página principal do filtrador"""
    return render_template('filtrador_combinacoes.html')


@filtrador_bp.route('/api/verificar-arquivo', methods=['GET'])
def verificar_arquivo():
    """Verifica se o arquivo de combinações existe"""
    try:
        caminho = FiltradorCombinacoesService.ARQUIVO_COMBINACOES
        existe = os.path.exists(caminho)

        if existe:
            tamanho_mb = os.path.getsize(caminho) / (1024 * 1024)
            # Conta linhas rapidamente
            with open(caminho, 'r', encoding='utf-8') as f:
                total_linhas = sum(1 for linha in f if linha.strip() and not linha.startswith('#'))

            return jsonify({
                'sucesso': True,
                'existe': True,
                'caminho': caminho,
                'tamanho_mb': round(tamanho_mb, 2),
                'total_combinacoes': total_linhas
            })
        else:
            return jsonify({
                'sucesso': True,
                'existe': False,
                'caminho': caminho,
                'mensagem': 'Arquivo ainda não foi gerado. Clique em "Gerar Arquivo" primeiro.'
            })

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/gerar-arquivo', methods=['POST'])
def gerar_arquivo():
    """Gera o arquivo .txt com todas as combinações"""
    try:
        resultado = gerar_arquivo_combinacoes_txt()
        return jsonify(resultado)

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/filtrar', methods=['POST'])
def filtrar_combinacoes():
    """
    Aplica filtros às combinações

    Body JSON exemplo:
    {
        "modo": "dinamico",  // ou "manual"
        "filtros": {
            "pares_impares": {"min_pares": 3, "max_pares": 4},
            "faixas_numeros": {"baixos": {"min": 2, "max": 2}, ...},
            "numeros_quentes": {"min_presentes": 3},
            "soma": {"min": 100, "max": 150},
            "sequencias": {"max_consecutivos": 2}
        },
        "limite_visualizacao": 1000  // apenas para exibição na tela
    }
    """
    try:
        dados = request.get_json(silent=True) or {}

        # Marca processamento como em andamento
        estado_processamento['em_andamento'] = True
        estado_processamento['pausado'] = False
        estado_processamento['progresso'] = 0

        # Extrai filtros (SEM limite - processa TODAS)
        filtros = dados.get('filtros', {})

        # Instancia o serviço
        servico = obter_servico()
        
        # Aplica os filtros (retorna TODAS as combinações filtradas)
        resultado = servico.aplicar_filtros(filtros)

        # Armazena resultado COMPLETO para downloads
        estado_processamento['resultado_atual'] = resultado

        # Para a resposta da API, limita apenas a visualização
        limite_visualizacao = dados.get('limite_visualizacao', 1000)
        resultado_visualizacao = resultado.copy()
        resultado_visualizacao['combinacoes_visualizacao'] = resultado['combinacoes'][:limite_visualizacao]
        resultado_visualizacao['total_retornado'] = len(resultado['combinacoes'][:limite_visualizacao])

        estado_processamento['em_andamento'] = False
        estado_processamento['progresso'] = 100

        return jsonify(resultado_visualizacao)

    except FileNotFoundError as e:
        estado_processamento['em_andamento'] = False
        return jsonify({
            'sucesso': False,
            'erro': str(e),
            'tipo_erro': 'arquivo_nao_encontrado'
        }), 404

    except Exception as e:
        estado_processamento['em_andamento'] = False
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/pausar', methods=['POST'])
def pausar_processamento():
    """Pausa o processamento atual"""
    try:
        estado_processamento['pausado'] = True
        return jsonify({
            'sucesso': True,
            'mensagem': 'Processamento pausado'
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/continuar', methods=['POST'])
def continuar_processamento():
    """Continua o processamento pausado"""
    try:
        estado_processamento['pausado'] = False
        return jsonify({
            'sucesso': True,
            'mensagem': 'Processamento retomado'
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/resetar', methods=['POST'])
def resetar_processamento():
    """Reseta o estado do processamento"""
    try:
        estado_processamento['em_andamento'] = False
        estado_processamento['pausado'] = False
        estado_processamento['progresso'] = 0
        estado_processamento['resultado_atual'] = None

        return jsonify({
            'sucesso': True,
            'mensagem': 'Estado resetado com sucesso'
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/status', methods=['GET'])
def obter_status():
    """Retorna o status atual do processamento"""
    try:
        return jsonify({
            'sucesso': True,
            'em_andamento': estado_processamento['em_andamento'],
            'pausado': estado_processamento['pausado'],
            'progresso': estado_processamento['progresso']
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/numeros-quentes-frios', methods=['GET'])
def obter_numeros_quentes_frios():
    """Retorna os números quentes e frios carregados do banco de dados"""
    try:
        # Força recarregamento se necessário
        servico_temp = FiltradorCombinacoesService()

        return jsonify({
            'sucesso': True,
            'numeros_quentes': FiltradorCombinacoesService.NUMEROS_QUENTES,
            'numeros_frios': FiltradorCombinacoesService.NUMEROS_FRIOS
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e),
            'numeros_quentes': [],
            'numeros_frios': []
        }), 500


@filtrador_bp.route('/api/analises-reais', methods=['GET'])
def obter_analises_reais():
    """Retorna TODAS as análises reais carregadas do banco de dados"""
    try:
        # Força recarregamento se necessário
        servico_temp = FiltradorCombinacoesService()

        return jsonify({
            'sucesso': True,
            'numeros_quentes': FiltradorCombinacoesService.NUMEROS_QUENTES,
            'numeros_frios': FiltradorCombinacoesService.NUMEROS_FRIOS,
            'padrao_pares': FiltradorCombinacoesService.PADRAO_PARES_MAIS_COMUM,
            'media_pares': FiltradorCombinacoesService.MEDIA_PARES,
            'soma_minima_ideal': FiltradorCombinacoesService.SOMA_MINIMA_IDEAL,
            'soma_maxima_ideal': FiltradorCombinacoesService.SOMA_MAXIMA_IDEAL,
            'soma_media': FiltradorCombinacoesService.SOMA_MEDIA,
            'padrao_faixas': FiltradorCombinacoesService.PADRAO_FAIXAS_MAIS_COMUM,
            'top_3_primos': FiltradorCombinacoesService.TOP_3_PADROES_PRIMOS,
            'top_3_digitos_iniciais': FiltradorCombinacoesService.TOP_3_PADROES_DIGITOS_INICIAIS
        })
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/obter-pagina', methods=['GET'])
def obter_pagina():
    """
    Retorna uma página específica das combinações filtradas

    Parâmetros:
        - pagina (int): Número da página (1-indexed)
        - por_pagina (int): Quantidade de itens por página (padrão: 100)
    """
    try:
        resultado = estado_processamento.get('resultado_atual')
        if not resultado:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum resultado disponível. Execute a filtragem primeiro.'
            }), 404

        # Parâmetros de paginação
        pagina = int(request.args.get('pagina', 1))
        por_pagina = int(request.args.get('por_pagina', 100))

        combinacoes = resultado.get('combinacoes', [])
        total = len(combinacoes)

        # Calcula total de páginas
        total_paginas = (total + por_pagina - 1) // por_pagina  # Arredonda para cima

        # Valida página
        if pagina < 1:
            pagina = 1
        if pagina > total_paginas:
            pagina = total_paginas

        # Calcula índices
        inicio = (pagina - 1) * por_pagina
        fim = min(inicio + por_pagina, total)

        # Extrai combinações da página
        combinacoes_pagina = combinacoes[inicio:fim]

        return jsonify({
            'sucesso': True,
            'pagina_atual': pagina,
            'total_paginas': total_paginas,
            'por_pagina': por_pagina,
            'total_combinacoes': total,
            'inicio': inicio + 1,  # 1-indexed para exibição
            'fim': fim,
            'combinacoes': combinacoes_pagina
        })

    except ValueError as e:
        return jsonify({
            'sucesso': False,
            'erro': 'Parâmetros inválidos. Use números inteiros para pagina e por_pagina.'
        }), 400

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/download/txt', methods=['GET'])
def download_txt():
    """Download das combinações filtradas em formato .txt com mês abreviado"""
    try:
        resultado = estado_processamento.get('resultado_atual')
        if not resultado:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum resultado disponível. Execute a filtragem primeiro.'
            }), 404

        combinacoes = resultado.get('combinacoes', [])
        mes = request.args.get('mes', 'Jan')  # Mês abreviado (Jan, Fev, Mar, etc.)

        # Gera conteúdo do arquivo
        conteudo = f"# Combinações Filtradas - Dia de Sorte\n"
        conteudo += f"# Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
        conteudo += f"# Mês da Sorte: {mes}\n"
        conteudo += f"# Total: {len(combinacoes)} combinações\n"
        estatisticas = resultado.get('estatisticas_filtros', [])
        if estatisticas:
            conteudo += "# Filtros aplicados e suas configurações detalhadas:\n"
            for stat in estatisticas:
                conteudo += f"# -> {stat.get('filtro')}: {stat.get('regra')}\n"
        else:
            conteudo += "# Filtros aplicados: Nenhum\n"
        conteudo += "\n"

        for comb in combinacoes:
            numeros = ','.join(f'{n:02d}' for n in comb)
            conteudo += f"{numeros} - {mes}\n"

        # Cria arquivo em memória
        buffer = io.BytesIO()
        buffer.write(conteudo.encode('utf-8'))
        buffer.seek(0)

        nome_personalizado = request.args.get('nome')
        import re
        if nome_personalizado:
            nome_limpo = re.sub(r'[\\/*?:"<>|]', "", nome_personalizado)
            # Acoplando o timestamp só se for modo manual para evitar substituições
            if "Manual" in nome_limpo or "Sem_Filtros" in nome_limpo:
                 nome_arquivo = f"{nome_limpo}_{datetime.now().strftime('%H%M%S')}.txt"
            else:
                 nome_arquivo = f"{nome_limpo}.txt"
        else:
            nome_arquivo = f"combinacoes_filtradas_{mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

        # ESPELHAMENTO INTERNO: Salvar Cópia Local na nova pasta
        try:
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            pasta_espelho = os.path.join(BASE_DIR, 'conferencia_filtros-baixados')
            os.makedirs(pasta_espelho, exist_ok=True)
            
            caminho_absoluto = os.path.join(pasta_espelho, nome_arquivo)
            
            # Evitar sobrescrever se o cara exportar duas vezes o mesmo nome
            contador = 1
            nome_base_esp = nome_arquivo.replace('.txt', '')
            while os.path.exists(caminho_absoluto):
                caminho_absoluto = os.path.join(pasta_espelho, f"{nome_base_esp} ({contador}).txt")
                contador += 1
                
            # Grava de fato
            print(f"[DEBUG] Espelhamento Manual em: {caminho_absoluto}")
            with open(caminho_absoluto, 'w', encoding='utf-8') as f:
                f.write(conteudo)
        except Exception as e_mirror:
            print(f"[ERRO ESPELHAMENTO] {e_mirror}")
            
        return send_file(
            buffer,
            mimetype='text/plain',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500

@filtrador_bp.route('/api/conferencia/injetar-atual', methods=['POST'])
def injetar_atual_conferencia():
    """Salva os resultados atuais na pasta de conferência sem exigir download manual"""
    print(f"[DEBUG] Rota injetar_atual_conferencia acionada!")
    try:
        resultado = estado_processamento.get('resultado_atual')
        if not resultado:
            print("[DEBUG] Nenhum resultado_atual encontrado no estado_processamento.")
            return jsonify({'sucesso': False, 'erro': 'Filtre primeiro antes de injetar.'}), 400
            
        combinacoes = resultado.get('combinacoes', [])
        print(f"[DEBUG] Injetando {len(combinacoes)} combinações...")
        mes = request.args.get('mes', 'Jan')
        
        # Gera o mesmo prefixo usado no download
        prefixo = request.args.get('nome', '[ABDUZIDO_FILTRADOR]')
        timestamp = datetime.now().strftime('%H%M%S')
        nome_arquivo = f"{prefixo}_{timestamp}.txt"
        
        # Conteúdo idêntico ao download
        conteudo = f"# Combinações Injetadas Automaticamente\n"
        conteudo += f"# Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
        conteudo += f"# Mês da Sorte: {mes}\n"
        conteudo += f"# Total: {len(combinacoes)} combinações\n\n"
        
        for comb in combinacoes:
            numeros = ','.join(f'{n:02d}' for n in comb)
            conteudo += f"{numeros} - {mes}\n"
            
        # Caminho Absoluto Robusto
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pasta_espelho = os.path.join(BASE_DIR, 'conferencia_filtros-baixados')
        os.makedirs(pasta_espelho, exist_ok=True)
        caminho = os.path.join(pasta_espelho, nome_arquivo)
        
        print(f"[DEBUG] Salvando em: {caminho}")
        with open(caminho, 'w', encoding='utf-8') as f:
            f.write(conteudo)
            
        return jsonify({'sucesso': True, 'arquivo': nome_arquivo})
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500

@filtrador_bp.route('/api/conferencia/limpar-pasta-espelho', methods=['POST'])
def limpar_pasta_espelho():
    """Exclui todos os arquivos da pasta de espelhamento de conferência"""
    try:
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pasta_espelho = os.path.join(BASE_DIR, 'conferencia_filtros-baixados')
        if os.path.exists(pasta_espelho):
            for file in os.listdir(pasta_espelho):
                if file.lower().endswith('.txt'):
                    path = os.path.join(pasta_espelho, file)
                    if os.path.isfile(path):
                        os.remove(path)
        return jsonify({'sucesso': True, 'mensagem': 'Matrizes da pasta espelho limpas.'})
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@filtrador_bp.route('/api/download/xlsx', methods=['GET'])
def download_xlsx():
    """Download das combinações filtradas em formato Excel (.xlsx)"""
    try:
        resultado = estado_processamento.get('resultado_atual')
        if not resultado:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum resultado disponível. Execute a filtragem primeiro.'
            }), 404

        combinacoes = resultado.get('combinacoes', [])
        mes = request.args.get('mes', 'Jan')

        # Importa openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return jsonify({
                'sucesso': False,
                'erro': 'Biblioteca openpyxl não instalada. Execute: pip install openpyxl'
            }), 500

        # Cria workbook
        wb = Workbook()
        
        # Cores do Dia de Sorte
        cor_verde = PatternFill(start_color="00A859", end_color="00A859", fill_type="solid")
        cor_amarelo = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # 1. ABA RESUMO_FILTROS
        ws_resumo = wb.active
        ws_resumo.title = "Resumo_Filtros"
        
        headers_resumo = ['Filtro', 'Descrição', 'Vantagem/Regra Aplicada']
        ws_resumo.append(headers_resumo)
        for cell in ws_resumo[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = cor_verde
            cell.alignment = Alignment(horizontal='center')
        
        estatisticas = resultado.get('estatisticas_filtros', [])
        for stat in estatisticas:
            nome_f = stat.get('filtro', '')
            desc = "Controle distribuição"
            vantagem = stat.get('regra', '')
            if 'Pares' in nome_f:
                desc = "Controla distribuição numérica"
            elif 'Primos' in nome_f:
                desc = "Equilíbrio de primos históricos"
            elif 'Soma' in nome_f:
                desc = "Evita somatórias absurdas"
            elif 'Específicos' in nome_f:
                desc = "Sniper de dezenas Frias e Quentes"
            elif 'Faixas' in nome_f:
                desc = "Distribui os quadrantes uniformemente"
            
            ws_resumo.append([nome_f, desc, vantagem])
            
        for col in ws_resumo.columns:
            ws_resumo.column_dimensions[col[0].column_letter].width = 30
            
        # 2. ABA REDUCAO_FILTROS
        ws_reducao = wb.create_sheet(title="Reducao_Filtros")
        ws_reducao.append(['Filtro', 'Antes (universo)', 'Depois', 'Redução (%)'])
        for cell in ws_reducao[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = cor_verde
            cell.alignment = Alignment(horizontal='center')
        
        for stat in estatisticas:
            pct_reduzido = 100 - stat.get('percentual', 100) # stats percentual é de sobrevivencia, entao 100 - % é redução?
            # actually o codigo original tem: 'percentual': round((len(resultado) / len(combinacoes)) * 100, 2)
            # entao % redução = 100 - X
            sobra_pct = stat.get('percentual', 100)
            reducao_pct = round(100 - sobra_pct, 2)
            
            ws_reducao.append([
                stat.get('filtro', ''),
                stat.get('antes', 0),
                stat.get('depois', 0),
                f"~{reducao_pct}%"
            ])
            
        for col in ws_reducao.columns:
            ws_reducao.column_dimensions[col[0].column_letter].width = 20

        # 3. ABA COMBINACOES
        ws_comb = wb.create_sheet(title=f"Filtradas {mes}")

        # Cabeçalho combinations
        headers = ['#', 'Num 1', 'Num 2', 'Num 3', 'Num 4', 'Num 5', 'Num 6', 'Num 7', 'Mês da Sorte', 'Soma']
        ws_comb.append(headers)

        # Estiliza cabeçalho
        for cell in ws_comb[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = cor_verde
            cell.alignment = Alignment(horizontal='center')

        # Dados
        for idx, comb in enumerate(combinacoes, 1):
            soma = sum(comb)
            linha = [idx] + list(comb) + [mes, soma]
            ws_comb.append(linha)

            # Alterna cores
            fill = cor_amarelo if idx % 2 == 0 else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for cell in ws_comb[idx + 1]:
                cell.alignment = Alignment(horizontal='center')
                if idx % 2 == 0 and cell.column > 1:  # Não aplica amarelo na primeira coluna
                    cell.fill = fill

        # Ajusta largura das colunas
        for col in ws_comb.columns:
            ws_comb.column_dimensions[col[0].column_letter].width = 12

        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        nome_arquivo = f"combinacoes_filtradas_{mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500


@filtrador_bp.route('/api/download/html', methods=['GET'])
def download_html():
    """Download das combinações filtradas em formato HTML completo"""
    try:
        resultado = estado_processamento.get('resultado_atual')
        if not resultado:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum resultado disponível. Execute a filtragem primeiro.'
            }), 404

        combinacoes = resultado.get('combinacoes', [])
        mes = request.args.get('mes', 'Jan')
        estatisticas = resultado.get('estatisticas_filtros', [])

        # Gera HTML completo
        html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Combinações Filtradas - Dia de Sorte</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #00A859 0%, #FFD700 100%);
            padding: 20px;
            margin: 0;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.3);
        }}
        h1 {{
            color: #00A859;
            text-align: center;
            margin-bottom: 10px;
        }}
        .info {{
            text-align: center;
            color: #666;
            margin-bottom: 30px;
        }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #00A859, #00D068);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }}
        .stat-card h3 {{
            margin: 0 0 10px 0;
            font-size: 2em;
        }}
        .stat-card p {{
            margin: 0;
            opacity: 0.9;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background: #00A859;
            color: white;
            padding: 12px;
            text-align: center;
            font-weight: bold;
        }}
        td {{
            padding: 10px;
            text-align: center;
            border-bottom: 1px solid #eee;
        }}
        tr:nth-child(even) {{
            background: #fffef0;
        }}
        tr:hover {{
            background: #FFD700;
            transition: background 0.3s;
        }}
        .numero {{
            display: inline-block;
            width: 35px;
            height: 35px;
            line-height: 35px;
            background: #00A859;
            color: white;
            border-radius: 50%;
            margin: 2px;
            font-weight: bold;
        }}
        .mes {{
            background: #FF8C00;
            color: white;
            padding: 5px 10px;
            border-radius: 5px;
            font-weight: bold;
        }}
        .filtros {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 10px;
            margin-bottom: 30px;
        }}
        .filtros h3 {{
            color: #00A859;
            margin-top: 0;
        }}
        .filtro-item {{
            background: white;
            padding: 10px;
            margin: 5px 0;
            border-left: 4px solid #00A859;
            border-radius: 5px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🍀 Dia de Sorte - Combinações Filtradas 🍀</h1>
        <div class="info">
            <p><strong>Data de Geração:</strong> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
            <p><strong>Mês da Sorte:</strong> <span class="mes">{mes}</span></p>
        </div>

        <div class="stats">
            <div class="stat-card">
                <h3>{resultado['total_original']:,}</h3>
                <p>Combinações Originais</p>
            </div>
            <div class="stat-card">
                <h3>{resultado['total_filtrado']:,}</h3>
                <p>Após Filtros</p>
            </div>
            <div class="stat-card">
                <h3>{resultado['percentual_reducao']}%</h3>
                <p>Redução</p>
            </div>
            <div class="stat-card">
                <h3>{resultado['tempo_processamento']}s</h3>
                <p>Tempo de Processamento</p>
            </div>
        </div>

        <div class="filtros">
            <h3>📊 Filtros Aplicados</h3>
"""

        for stat in estatisticas:
            html += f"""
            <div class="filtro-item">
                <strong>{stat['filtro']}:</strong> {stat['regra']}<br>
                <small>Eliminadas: {stat['eliminadas']:,} ({100 - stat['percentual']:.1f}%)</small>
            </div>
"""

        html += """
        </div>

        <h2 style="color: #00A859; text-align: center;">🎲 Combinações Selecionadas</h2>
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th colspan="7">Números</th>
                    <th>Soma</th>
                </tr>
            </thead>
            <tbody>
"""

        for idx, comb in enumerate(combinacoes, 1):
            soma = sum(comb)
            numeros_html = ''.join(f'<span class="numero">{n:02d}</span>' for n in comb)
            html += f"""
                <tr>
                    <td><strong>{idx}</strong></td>
                    <td colspan="7">{numeros_html}</td>
                    <td><strong>{soma}</strong></td>
                </tr>
"""

        html += """
            </tbody>
        </table>
    </div>
</body>
</html>
"""

        # Cria arquivo em memória
        buffer = io.BytesIO()
        buffer.write(html.encode('utf-8'))
        buffer.seek(0)

        nome_arquivo = f"combinacoes_filtradas_{mes}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"

        return send_file(
            buffer,
            mimetype='text/html',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': str(e)
        }), 500
